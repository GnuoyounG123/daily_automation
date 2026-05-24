#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import asyncio
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter.constants import *
from tkinter.scrolledtext import ScrolledText
import json
import csv
import re
import threading
import subprocess
import sys
import webbrowser
from pathlib import Path
from datetime import datetime
from io import StringIO

from .config_manager import ConfigManager, DEFAULT_NEWS_SOURCES, SOURCE_CATEGORIES, DEFAULT_TIME_SLOTS

try:
    import ttkbootstrap as ttk
    from ttkbootstrap.constants import *
    TTKBOOTSTRAP_AVAILABLE = True
except ModuleNotFoundError:
    import tkinter.ttk as ttk
    TTKBOOTSTRAP_AVAILABLE = False
    PRIMARY = "primary"
    SECONDARY = "secondary"
    SUCCESS = "success"
    INFO = "info"
    WARNING = "warning"
    DANGER = "danger"
    OUTLINE = "outline"


POPULAR_SCHOOLS = [
    "浙江大学", "清华大学", "北京大学", "复旦大学", "上海交通大学",
    "南京大学", "中国科学技术大学", "武汉大学", "华中科技大学", "中山大学",
    "同济大学", "东南大学", "哈尔滨工业大学", "西安交通大学", "北京航空航天大学",
    "四川大学", "南开大学", "天津大学", "厦门大学", "中国人民大学",
    "其他（手动输入）"
]

GRADES = ["大一", "大二", "大三", "大四", "研一", "研二", "研三", "博一", "博二", "博三"]

INTEREST_AREAS = [
    ("🤖 人工智能", ["artificial intelligence", "machine learning", "deep learning", "NLP", "computer vision"], ["人工智能", "机器学习", "深度学习"]),
    ("📊 数据科学", ["data science", "big data", "data mining", "statistics", "data analysis"], ["数据科学", "大数据", "数据挖掘"]),
    ("💻 计算机系统", ["operating systems", "distributed systems", "cloud computing", "database"], ["计算机系统", "分布式系统", "云计算"]),
    ("🔒 网络安全", ["cybersecurity", "cryptography", "network security", "privacy"], ["网络安全", "密码学"]),
    ("💰 经济学", ["economics", "econometrics", "macroeconomics", "microeconomics"], ["经济学", "计量经济学"]),
    ("📈 金融学", ["finance", "financial engineering", "quantitative finance", "fintech"], ["金融学", "金融工程"]),
    ("🏛️ 公共管理", ["public governance", "public policy", "digital governance", "e-government"], ["公共管理", "数字治理"]),
    ("🧠 心理学", ["psychology", "cognitive science", "behavioral science"], ["心理学", "认知科学"]),
    ("🏥 医学", ["medicine", "biomedical", "clinical research", "public health"], ["医学", "生物医学"]),
    ("📚 教育学", ["education", "pedagogy", "educational technology"], ["教育学", "教育技术"]),
    ("⚖️ 法学", ["law", "jurisprudence", "constitutional law"], ["法学", "法律"]),
    ("✍️ 文学", ["literature", "linguistics", "cultural studies"], ["文学", "语言学"]),
    ("🌱 环境科学", ["environmental science", "climate change", "sustainability"], ["环境科学", "气候变化"]),
    ("🔬 材料科学", ["materials science", "nanotechnology", "polymer science"], ["材料科学", "纳米技术"]),
]

DAY_NAMES_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
DAY_NAMES_EN = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DAY_MAP = dict(zip(DAY_NAMES_CN, DAY_NAMES_EN))


def _get_app_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return _get_project_root() / "runtime_local"


def _get_source_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _get_project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _decode_pwd(value: str) -> str:
    import base64
    if not value:
        return ""
    if value.startswith("fernet:"):
        try:
            from .password_crypto import PasswordCrypto
            crypto = PasswordCrypto(_get_app_dir())
            return crypto.decrypt(value)
        except Exception:
            return value
    if value.startswith("enc:"):
        try:
            return base64.b64decode(value[4:]).decode('utf-8')
        except Exception:
            return value
    return value


def _encode_pwd(value: str) -> str:
    import base64
    if value and not value.startswith(("enc:", "fernet:")):
        try:
            from .password_crypto import PasswordCrypto
            return PasswordCrypto(_get_app_dir()).encrypt(value)
        except Exception:
            return "enc:" + base64.b64encode(value.encode('utf-8')).decode('utf-8')
    return value


def parse_course_table(text):
    courses = []
    lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
    if not lines:
        return courses

    separator = None
    for line in lines[:5]:
        if '\t' in line:
            separator = '\t'
            break
        elif '|' in line:
            separator = '|'
            break
        elif ',' in line and line.count(',') >= 2:
            separator = ','
            break

    rows = []
    for line in lines:
        if separator:
            parts = [p.strip() for p in line.split(separator)]
        else:
            parts = [p.strip() for p in re.split(r'\s{2,}', line)]
        if len(parts) >= 2:
            non_empty = sum(1 for p in parts if p)
            if non_empty <= 2 and len(parts) > 4:
                continue
            rows.append(parts)

    if not rows:
        return courses

    name_col = -1
    time_col = -1
    location_col = -1
    teacher_col = -1
    day_col = -1

    header_idx = -1
    for row_idx, row in enumerate(rows):
        temp_name = temp_time = temp_loc = temp_teacher = temp_day = -1
        found = False
        for i, cell in enumerate(row):
            cl = cell.lower()
            if any(k in cl for k in ['课程名称', '课程名', 'name', 'course name']):
                temp_name = i; found = True
            elif any(k in cl for k in ['上课时间', 'time', '节次']):
                temp_time = i; found = True
            elif any(k in cl for k in ['上课地点', 'location', 'room', '地点', '教室']):
                temp_loc = i; found = True
            elif any(k in cl for k in ['教师姓名', '任课教师', 'teacher', '授课教师']):
                temp_teacher = i; found = True
            elif any(k in cl for k in ['星期', 'day', 'week']) and not _looks_like_time_or_day(cell):
                temp_day = i; found = True
        if found and (temp_name >= 0 or temp_time >= 0):
            name_col, time_col, location_col, teacher_col, day_col = temp_name, temp_time, temp_loc, temp_teacher, temp_day
            header_idx = row_idx
            break

    if header_idx < 0:
        first_row = rows[0]
        for i, cell in enumerate(first_row):
            cl = cell.lower()
            if any(k in cl for k in ['课程名', '名称', 'course', 'name', '科目']):
                name_col = i
            elif any(k in cl for k in ['时间', 'time', '节次']):
                time_col = i
            elif any(k in cl for k in ['地点', '教室', 'location', 'room']):
                location_col = i
            elif any(k in cl for k in ['教师', 'teacher', '授课']):
                teacher_col = i
        if name_col < 0 and time_col < 0:
            if len(first_row) >= 4:
                if _looks_like_time_or_day(first_row[0]):
                    time_col, name_col, teacher_col, location_col = 0, 1, 2, 3
                else:
                    name_col, time_col, location_col, teacher_col = 0, 1, 2, 3
            elif len(first_row) >= 3:
                if _looks_like_time_or_day(first_row[0]):
                    time_col, name_col, location_col = 0, 1, 2
                else:
                    name_col, time_col, location_col = 0, 1, 2
            elif len(first_row) >= 2:
                if _looks_like_time_or_day(first_row[0]):
                    time_col, name_col = 0, 1
                else:
                    name_col, time_col = 0, 1
        header_idx = 0

    data_start = header_idx + 1

    for row in rows[data_start:]:
        if len(row) <= max(name_col, 0):
            continue

        name = row[name_col] if name_col >= 0 and name_col < len(row) else ""
        time_val = row[time_col] if time_col >= 0 and time_col < len(row) else ""
        location = row[location_col] if location_col >= 0 and location_col < len(row) else ""
        teacher = row[teacher_col] if teacher_col >= 0 and teacher_col < len(row) else ""
        day_str = row[day_col] if day_col >= 0 and day_col < len(row) else ""

        if not name:
            continue

        time_parts = [t.strip() for t in time_val.split(';') if t.strip()] if time_val else [""]
        loc_parts = [l.strip() for l in location.split(';') if l.strip()] if location else [""]

        for ti, tp in enumerate(time_parts):
            day = ""
            if day_str:
                day = _parse_day(day_str)
            if not day and tp:
                day = _parse_day(tp)
            if not day:
                day = "Monday"

            clean_time = _strip_day_from_time(tp) if tp and day else tp

            loc = loc_parts[ti] if ti < len(loc_parts) else (loc_parts[-1] if loc_parts else "")

            courses.append({
                'day': day,
                'name': name,
                'time': clean_time,
                'location': loc,
                'teacher': teacher,
                'note': ''
            })

    return courses


def _looks_like_time_or_day(text):
    if not text:
        return False
    if re.search(r'周[一二三四五六日末1-7]', text):
        return True
    if re.search(r'星期[一二三四五六日]', text):
        return True
    if re.search(r'(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)', text, re.IGNORECASE):
        return True
    if re.search(r'^[0-9]+-[0-9]+节', text):
        return True
    if re.search(r'^\d{1,2}:\d{2}', text):
        return True
    return False


def _parse_day(text):
    mapping = {
        '周一': 'Monday', '周二': 'Tuesday', '周三': 'Wednesday',
        '周四': 'Thursday', '周五': 'Friday', '周六': 'Saturday', '周日': 'Sunday',
        '星期一': 'Monday', '星期二': 'Tuesday', '星期三': 'Wednesday',
        '星期四': 'Thursday', '星期五': 'Friday', '星期六': 'Saturday', '星期日': 'Sunday',
        '周1': 'Monday', '周2': 'Tuesday', '周3': 'Wednesday',
        '周4': 'Thursday', '周5': 'Friday', '周6': 'Saturday', '周7': 'Sunday',
    }
    for cn, en in mapping.items():
        if cn in text:
            return en

    en_mapping = {
        'monday': 'Monday', 'tuesday': 'Tuesday', 'wednesday': 'Wednesday',
        'thursday': 'Thursday', 'friday': 'Friday', 'saturday': 'Saturday', 'sunday': 'Sunday',
    }
    text_lower = text.lower()
    for en_key, en_val in en_mapping.items():
        if en_key in text_lower:
            return en_val

    return ""


def _strip_day_from_time(text):
    text = re.sub(r'周[一二三四五六日末1-7]', '', text)
    text = re.sub(r'星期[一二三四五六日]', '', text)
    text = re.sub(r'(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)', '', text, flags=re.IGNORECASE)
    return text.strip()


class DailyAutomationApp:

    def __init__(self, root):
        self.root = root
        self.root.geometry("1000x750")
        self.root.minsize(900, 650)

        self.app_dir = _get_app_dir()
        self.config_manager = ConfigManager(self.app_dir)

        is_first_run = self.config_manager.ensure_default_config()
        self.config_manager.ensure_default_schedule()

        self.create_ui()
        self.load_all_config()
        if is_first_run:
            self.root.after(100, self.show_setup_wizard)

    def _safe_showerror(self, title, message):
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(lambda: self.root.attributes('-topmost', False))
        messagebox.showerror(title, message)

    def show_setup_wizard(self):
        wizard = SetupWizard(self.root, self.config_manager)
        if wizard.completed:
            self.load_all_config()

    def create_ui(self):
        self.notebook = ttk.Notebook(self.root, bootstyle=PRIMARY)
        self.notebook.pack(fill='both', expand=True, padx=15, pady=15)

        self.create_home_tab()
        self.create_control_tab()
        self.create_sources_tab()
        self.create_keywords_tab()
        self.create_reminders_tab()
        self.create_email_tab()
        self.create_api_keys_tab()
        self.create_schedule_tab()
        self.create_about_tab()

        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief='sunken', padding=5, bootstyle=SECONDARY)
        status_bar.pack(fill='x', side='bottom')

    # ========== 首页 ==========
    def create_home_tab(self):
        frame = ttk.Frame(self.notebook, padding=18)
        self.notebook.add(frame, text="🏠 首页")

        header = ttk.Frame(frame)
        header.pack(fill='x', pady=(4, 14))
        ttk.Label(header, text="今日简报", font=('Microsoft YaHei', 22, 'bold'), bootstyle=PRIMARY).pack(side='left')
        ttk.Button(header, text="生成并发送", command=self.run_once, bootstyle=PRIMARY, width=14).pack(side='right', padx=(8, 0))
        ttk.Button(header, text="进入控制台", command=self.open_control_tab, bootstyle=INFO, width=12).pack(side='right')

        summary = ttk.Labelframe(frame, text="今日状态", padding=12)
        summary.pack(fill='x', pady=(0, 12))

        self.home_status_labels = {}
        stats = [
            ('last_report', '最新简报', '未生成'),
            ('news_count', '新闻源', '0'),
            ('email_status', '邮件', '未启用'),
            ('last_issue', '运行问题', '待检查'),
        ]

        for key, label, default in stats:
            col = ttk.Frame(summary)
            col.pack(side='left', expand=True, fill='x', padx=8)
            ttk.Label(col, text=label, font=('Microsoft YaHei', 10), bootstyle=SECONDARY).pack()
            self.home_status_labels[key] = ttk.Label(col, text=default, font=('Microsoft YaHei', 14, 'bold'), bootstyle=PRIMARY)
            self.home_status_labels[key].pack(pady=(4, 0))

        body = ttk.Frame(frame)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        briefing_frame = ttk.Labelframe(body, text="简报预览", padding=10)
        briefing_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 12))
        self.briefing_text = ScrolledText(briefing_frame, height=22, wrap='word', state='disabled', font=('Microsoft YaHei', 10))
        self.briefing_text.pack(fill='both', expand=True)

        side = ttk.Frame(body)
        side.grid(row=0, column=1, sticky='nsew')
        side.rowconfigure(0, weight=2)
        side.rowconfigure(1, weight=1)

        issue_frame = ttk.Labelframe(side, text="运行问题", padding=10)
        issue_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 12))
        self.issue_text = ScrolledText(issue_frame, height=12, wrap='word', state='disabled', font=('Microsoft YaHei', 10))
        self.issue_text.pack(fill='both', expand=True)

        quick_frame = ttk.Labelframe(side, text="快捷入口", padding=10)
        quick_frame.grid(row=1, column=0, sticky='nsew')
        ttk.Button(quick_frame, text="打开数据目录", command=self.open_data_dir, bootstyle=INFO).pack(fill='x', pady=4)
        ttk.Button(quick_frame, text="邮件设置", command=lambda: self.notebook.select(5), bootstyle=OUTLINE).pack(fill='x', pady=4)
        ttk.Button(quick_frame, text="新闻源设置", command=lambda: self.notebook.select(2), bootstyle=OUTLINE).pack(fill='x', pady=4)
        ttk.Button(quick_frame, text="API 密钥", command=lambda: self.notebook.select(6), bootstyle=OUTLINE).pack(fill='x', pady=4)

    # ========== 控制台 ==========
    def create_control_tab(self):
        frame = ttk.Frame(self.notebook, padding=18)
        self.notebook.add(frame, text="🧭 控制台")

        ttk.Label(frame, text="运行控制台", font=('Microsoft YaHei', 20, 'bold'), bootstyle=PRIMARY).pack(anchor='w', pady=(4, 14))

        card_frame = ttk.Labelframe(frame, text="配置概览", padding=15)
        card_frame.pack(fill='x', pady=(0, 12))

        self.status_labels = {}
        stats = [('news_count', '新闻源', '0'), ('keyword_count', '关键词', '0'),
                 ('reminder_count', '提醒', '0'), ('email_status', '邮件', '未配置')]

        for key, label, default in stats:
            col = ttk.Frame(card_frame)
            col.pack(side='left', expand=True, padx=10, pady=5)
            ttk.Label(col, text=label, font=('Microsoft YaHei', 10), bootstyle=SECONDARY).pack()
            self.status_labels[key] = ttk.Label(col, text=default, font=('Microsoft YaHei', 18, 'bold'), bootstyle=PRIMARY)
            self.status_labels[key].pack(pady=(5, 0))

        action_frame = ttk.Labelframe(frame, text="快速操作", padding=15)
        action_frame.pack(fill='x', pady=(0, 12))

        btn_frame = ttk.Frame(action_frame)
        btn_frame.pack(pady=8)
        ttk.Button(btn_frame, text="一键生成今日简报", command=self.run_once, width=22, bootstyle=PRIMARY).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="异步生成(实验)", command=self.run_once_async, width=18, bootstyle=SUCCESS).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="只检查提醒", command=self.run_remind, width=14, bootstyle=INFO).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="打开数据目录", command=self.open_data_dir, width=16, bootstyle=OUTLINE).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="导出配置", command=self.export_config, width=12, bootstyle=OUTLINE).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="导入配置", command=self.import_config, width=12, bootstyle=OUTLINE).pack(side='left', padx=8)

        log_frame = ttk.Labelframe(frame, text="最近运行日志", padding=10)
        log_frame.pack(fill='both', expand=True)

        self.log_text = ScrolledText(log_frame, height=12, state='disabled', font=('Consolas', 10))
        self.log_text.pack(fill='both', expand=True, padx=5, pady=5)

    def open_control_tab(self):
        self.notebook.select(1)

    def refresh_home_status(self):
        config = self.config_manager.get_config()

        sources = config.get('news_sources', [])
        enabled = len([s for s in sources if s.get('enabled', True)])
        self._set_status_value('news_count', f"{enabled}/{len(sources)}")

        keywords = config.get('keywords', []) + config.get('keywords_cn', [])
        self._set_status_value('keyword_count', str(len(keywords)))

        reminders = config.get('reminders', [])
        self._set_status_value('reminder_count', str(len(reminders)))

        email = config.get('email', {})
        self._set_status_value('email_status', "已启用" if email.get('enabled') else "未启用")

        report_path = self.load_latest_briefing()
        self._set_status_value('last_report', report_path.name if report_path else "未生成")
        issues = self.collect_runtime_issues(config, report_path)
        self._set_status_value('last_issue', "正常" if not issues else f"{len(issues)}项")
        self._set_text_widget(self.issue_text, "\n".join(f"• {item}" for item in issues) if issues else "暂无需要处理的问题。")
        self.load_recent_logs()

    def _set_status_value(self, key, value):
        for attr in ('status_labels', 'home_status_labels'):
            labels = getattr(self, attr, {})
            if key in labels:
                labels[key].config(text=value)

    def _set_text_widget(self, widget, value):
        widget.config(state='normal')
        widget.delete(1.0, tk.END)
        widget.insert(tk.END, value)
        widget.config(state='disabled')

    def load_latest_briefing(self):
        data_dir = self.app_dir / "data"
        report_files = sorted(data_dir.glob("academic_briefing_*.md"), key=lambda x: x.stat().st_mtime, reverse=True) if data_dir.exists() else []
        if not report_files:
            self._set_text_widget(self.briefing_text, "还没有生成过简报。点击右上角“生成并发送”开始第一次运行。")
            return None

        latest = report_files[0]
        try:
            content = latest.read_text(encoding='utf-8', errors='replace')
        except Exception as e:
            self._set_text_widget(self.briefing_text, f"简报读取失败：{e}")
            return latest

        self._set_text_widget(self.briefing_text, content.strip() or "最新简报为空。")
        return latest

    def collect_runtime_issues(self, config, report_path):
        issues = []
        if not report_path:
            issues.append("今日简报尚未生成")

        enabled_sources = [s for s in config.get('news_sources', []) if s.get('enabled', True)]
        if not enabled_sources:
            issues.append("没有启用的新闻源")

        email = config.get('email', {})
        if not email.get('enabled'):
            issues.append("邮件发送未启用")
        elif not email.get('sender_email') or not email.get('sender_password') or not email.get('receiver_email'):
            issues.append("邮件配置不完整")

        recent_problem = self.get_recent_problem_line()
        if recent_problem:
            issues.append(recent_problem)
        return issues

    def get_recent_problem_line(self):
        log_dir = self.app_dir / "logs"
        if not log_dir.exists():
            return None

        patterns = ('ERROR', 'WARNING', '失败', '异常', 'failed', 'Failed')
        for log_file in sorted(log_dir.glob("*.log"), key=lambda x: x.stat().st_mtime, reverse=True)[:3]:
            try:
                lines = log_file.read_text(encoding='utf-8', errors='replace').splitlines()
            except Exception:
                continue
            for line in reversed(lines[-80:]):
                if any(pattern in line for pattern in patterns):
                    return f"最近日志提示：{line[-140:]}"
        return None

    def notify_task_issue(self, title, message, severity="warning"):
        self.status_var.set(message)
        self.refresh_home_status()
        if severity == "error":
            self._safe_showerror(title, message)
        else:
            self.root.lift()
            self.root.attributes('-topmost', True)
            self.root.after_idle(lambda: self.root.attributes('-topmost', False))
            messagebox.showwarning(title, message)

    def load_recent_logs(self):
        if not hasattr(self, 'log_text'):
            return
        log_dir = self.app_dir / "logs"
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)

        if log_dir.exists():
            log_files = sorted(log_dir.glob("*.log"), key=lambda x: x.name, reverse=True)[:3]
            for log_file in log_files:
                try:
                    with open(log_file, 'r', encoding='utf-8') as f:
                        lines = f.readlines()[-20:]
                        self.log_text.insert(tk.END, f"=== {log_file.name} ===" + "\n")
                        self.log_text.insert(tk.END, ''.join(lines) + "\n\n")
                except Exception:
                    pass
        else:
            self.log_text.insert(tk.END, "暂无运行记录")

        self.log_text.config(state='disabled')

    # ========== 新闻源管理 ==========
    def create_sources_tab(self):
        frame = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(frame, text="📰 新闻源")

        list_frame = ttk.Labelframe(frame, text="已配置的新闻源", padding=10)
        list_frame.pack(fill='both', expand=True, pady=5)

        columns = ('name', 'url', 'type', 'enabled')
        self.sources_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=10, bootstyle=PRIMARY)

        self.sources_tree.heading('name', text='名称')
        self.sources_tree.heading('url', text='URL')
        self.sources_tree.heading('type', text='类型')
        self.sources_tree.heading('enabled', text='状态')

        self.sources_tree.column('name', width=150)
        self.sources_tree.column('url', width=400)
        self.sources_tree.column('type', width=100)
        self.sources_tree.column('enabled', width=100, anchor='center')

        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.sources_tree.yview)
        self.sources_tree.configure(yscrollcommand=scrollbar.set)

        self.sources_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=15)

        ttk.Button(btn_frame, text="➕ 添加", command=self.add_source, bootstyle=SUCCESS).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✏️ 编辑", command=self.edit_source, bootstyle=WARNING).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🗑️ 删除", command=self.delete_source, bootstyle=DANGER).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🔄 切换状态", command=self.toggle_source, bootstyle=INFO).pack(side='left', padx=5)

    def load_sources(self):
        for item in self.sources_tree.get_children():
            self.sources_tree.delete(item)

        sources = self.config_manager.get_news_sources()
        for source in sources:
            enabled = "✅ 启用" if source.get('enabled', True) else "❌ 禁用"
            self.sources_tree.insert('', 'end', values=(
                source.get('name', ''),
                source.get('url', ''),
                source.get('type', 'rss'),
                enabled
            ))

    def add_source(self):
        dialog = SourceDialog(self.root, "添加新闻源")
        if dialog.result:
            self.config_manager.add_news_source(**dialog.result)
            self.load_sources()
            self.refresh_home_status()
            messagebox.showinfo("成功", "新闻源已添加")

    def edit_source(self):
        selected = self.sources_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要编辑的新闻源")
            return

        index = self.sources_tree.index(selected[0])
        sources = self.config_manager.get_news_sources()
        source = sources[index]

        dialog = SourceDialog(self.root, "编辑新闻源", source)
        if dialog.result:
            self.config_manager.update_news_source(index, **dialog.result)
            self.load_sources()
            messagebox.showinfo("成功", "新闻源已更新")

    def delete_source(self):
        selected = self.sources_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的新闻源")
            return

        if messagebox.askyesno("确认", "确定要删除此新闻源吗？"):
            index = self.sources_tree.index(selected[0])
            self.config_manager.delete_news_source(index)
            self.load_sources()
            self.refresh_home_status()

    def toggle_source(self):
        selected = self.sources_tree.selection()
        if not selected:
            return

        index = self.sources_tree.index(selected[0])
        sources = self.config_manager.get_news_sources()
        current = sources[index].get('enabled', True)
        self.config_manager.update_news_source(index, enabled=not current)
        self.load_sources()

    # ========== 关键词管理 ==========
    def create_keywords_tab(self):
        frame = ttk.Frame(self.notebook, padding=20)
        self.notebook.add(frame, text="🔑 关键词")

        ttk.Label(frame, text="设置用于筛选学术信息的关键词（每行一个）", font=('Microsoft YaHei', 11, 'bold')).pack(pady=(0, 15))

        col_frame = ttk.Frame(frame)
        col_frame.pack(fill='both', expand=True, pady=5)

        en_frame = ttk.Labelframe(col_frame, text="英文关键词", padding=10)
        en_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))

        self.en_keywords_text = ScrolledText(en_frame, width=40)
        self.en_keywords_text.pack(fill='both', expand=True)

        cn_frame = ttk.Labelframe(col_frame, text="中文关键词", padding=10)
        cn_frame.pack(side='left', fill='both', expand=True, padx=(10, 0))

        self.cn_keywords_text = ScrolledText(cn_frame, width=40)
        self.cn_keywords_text.pack(fill='both', expand=True)

        ttk.Button(frame, text="💾 保存关键词", command=self.save_keywords, bootstyle=SUCCESS, width=20).pack(pady=20)

        suggest_frame = ttk.Labelframe(frame, text="💡 关键词建议", padding=10)
        suggest_frame.pack(fill='x', pady=5)

        suggestions = "学术研究: artificial intelligence, machine learning, deep learning\n" \
                     "公共管理: public governance, digital governance, e-government\n" \
                     "数据科学: big data, data mining, algorithm"
        ttk.Label(suggest_frame, text=suggestions, justify='left', bootstyle=SECONDARY).pack(anchor='w')

    def load_keywords(self):
        keywords = self.config_manager.get_keywords()

        self.en_keywords_text.delete(1.0, tk.END)
        self.en_keywords_text.insert(tk.END, '\n'.join(keywords.get('keywords', [])))

        self.cn_keywords_text.delete(1.0, tk.END)
        self.cn_keywords_text.insert(tk.END, '\n'.join(keywords.get('keywords_cn', [])))

    def save_keywords(self):
        en_text = self.en_keywords_text.get(1.0, tk.END)
        cn_text = self.cn_keywords_text.get(1.0, tk.END)

        en_keywords = [k.strip() for k in en_text.split('\n') if k.strip()]
        cn_keywords = [k.strip() for k in cn_text.split('\n') if k.strip()]

        if self.config_manager.update_keywords(en_keywords, cn_keywords):
            messagebox.showinfo("成功", "关键词已保存")
            self.refresh_home_status()

    # ========== 提醒管理 ==========
    def create_reminders_tab(self):
        frame = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(frame, text="⏰ 提醒")

        list_frame = ttk.Labelframe(frame, text="已配置的提醒", padding=10)
        list_frame.pack(fill='both', expand=True, pady=5)

        columns = ('time', 'title', 'description')
        self.reminders_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=10, bootstyle=PRIMARY)

        self.reminders_tree.heading('time', text='时间')
        self.reminders_tree.heading('title', text='标题')
        self.reminders_tree.heading('description', text='描述')

        self.reminders_tree.column('time', width=100, anchor='center')
        self.reminders_tree.column('title', width=200)
        self.reminders_tree.column('description', width=500)

        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.reminders_tree.yview)
        self.reminders_tree.configure(yscrollcommand=scrollbar.set)

        self.reminders_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=15)

        ttk.Button(btn_frame, text="➕ 添加", command=self.add_reminder, bootstyle=SUCCESS).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✏️ 编辑", command=self.edit_reminder, bootstyle=WARNING).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🗑️ 删除", command=self.delete_reminder, bootstyle=DANGER).pack(side='left', padx=5)

    def load_reminders(self):
        for item in self.reminders_tree.get_children():
            self.reminders_tree.delete(item)

        reminders = self.config_manager.get_reminders()
        for r in reminders:
            self.reminders_tree.insert('', 'end', values=(
                r.get('time', ''),
                r.get('title', ''),
                r.get('description', '')
            ))

    def add_reminder(self):
        dialog = ReminderDialog(self.root, "添加提醒")
        if dialog.result:
            self.config_manager.add_reminder(**dialog.result)
            self.load_reminders()
            self.refresh_home_status()
            messagebox.showinfo("成功", "提醒已添加")

    def edit_reminder(self):
        selected = self.reminders_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要编辑的提醒")
            return

        index = self.reminders_tree.index(selected[0])
        reminders = self.config_manager.get_reminders()
        reminder = reminders[index]

        dialog = ReminderDialog(self.root, "编辑提醒", reminder)
        if dialog.result:
            self.config_manager.update_reminder(index, **dialog.result)
            self.load_reminders()
            messagebox.showinfo("成功", "提醒已更新")

    def delete_reminder(self):
        selected = self.reminders_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的提醒")
            return

        if messagebox.askyesno("确认", "确定要删除此提醒吗？"):
            index = self.reminders_tree.index(selected[0])
            self.config_manager.delete_reminder(index)
            self.load_reminders()
            self.refresh_home_status()

    # ========== 邮箱配置 ==========
    def create_email_tab(self):
        frame = ttk.Frame(self.notebook, padding=20)
        self.notebook.add(frame, text="📧 邮箱")

        self.email_enabled = tk.BooleanVar()
        # 优化成现代圆角 Toggle 开关
        ttk.Checkbutton(frame, text="启用邮件发送自动播报", variable=self.email_enabled, bootstyle="round-toggle").pack(pady=10, anchor='w')

        form_frame = ttk.Labelframe(frame, text="邮件服务器设置", padding=20)
        form_frame.pack(fill='x', pady=10)

        labels = ['SMTP服务器:', 'SMTP端口:', '发件人邮箱:', '授权密码:', '收件人邮箱:', '邮件主题前缀:', '天气城市:']
        self.email_entries = {}

        for i, label in enumerate(labels):
            row = ttk.Frame(form_frame)
            row.pack(fill='x', pady=8)

            ttk.Label(row, text=label, width=15, font=('Microsoft YaHei', 10)).pack(side='left')

            key = label.replace(':', '').replace(' ', '_')
            entry = ttk.Entry(row, width=50)
            if '密码' in label:
                entry.config(show='•')
            entry.pack(side='left', padx=15)
            self.email_entries[key] = entry

        ttk.Button(frame, text="💾 保存邮箱配置", command=self.save_email_config, bootstyle=SUCCESS, width=20).pack(pady=20)

        help_frame = ttk.Labelframe(frame, text="📋 常用邮箱SMTP配置", padding=15)
        help_frame.pack(fill='x', pady=10)

        help_text = """QQ邮箱: smtp.qq.com 端口587 (需要开启SMTP服务并获取授权码)
163邮箱: smtp.163.com 端口25 (需要开启SMTP服务)
Gmail: smtp.gmail.com 端口587 (需要应用专用密码)
Outlook: smtp-mail.outlook.com 端口587"""
        ttk.Label(help_frame, text=help_text, justify='left', bootstyle=SECONDARY).pack(anchor='w')

    def load_email_config(self):
        email = self.config_manager.get_email_config()
        self.email_enabled.set(email.get('enabled', False))
        for key in self.email_entries:
            self.email_entries[key].delete(0, tk.END)
        self.email_entries['SMTP服务器'].insert(0, email.get('smtp_server', 'smtp.qq.com'))
        self.email_entries['SMTP端口'].insert(0, str(email.get('smtp_port', 587)))
        self.email_entries['发件人邮箱'].insert(0, email.get('sender_email', ''))
        self.email_entries['授权密码'].insert(0, _decode_pwd(email.get('sender_password', '')))
        self.email_entries['收件人邮箱'].insert(0, email.get('receiver_email', ''))
        self.email_entries['邮件主题前缀'].insert(0, email.get('subject_prefix', '[学术简报]'))
        weather_city = self.config_manager.get_weather_city()
        self.email_entries['天气城市'].insert(0, weather_city)

    def save_email_config(self):
        port_str = self.email_entries['SMTP端口'].get().strip()
        try:
            port = int(port_str) if port_str else 587
        except ValueError:
            self._safe_showerror("错误", "SMTP端口必须为数字")
            return

        sender_email = self.email_entries['发件人邮箱'].get().strip()
        sender_pwd = self.email_entries['授权密码'].get().strip()
        receiver_email = self.email_entries['收件人邮箱'].get().strip()
        weather_city = self.email_entries['天气城市'].get().strip() or 'Beijing'

        auto_enabled = False
        if sender_email and sender_pwd and receiver_email and not self.email_enabled.get():
            self.email_enabled.set(True)
            auto_enabled = True

        config = {
            'enabled': self.email_enabled.get(),
            'smtp_server': self.email_entries['SMTP服务器'].get(),
            'smtp_port': port,
            'sender_email': sender_email,
            'sender_password': _encode_pwd(sender_pwd),
            'receiver_email': receiver_email,
            'subject_prefix': self.email_entries['邮件主题前缀'].get()
        }

        if self.config_manager.update_email_config(config):
            self.config_manager.update_weather_city(weather_city)
            msg = "邮箱配置已保存"
            if auto_enabled:
                msg += "\n\n✅ 已自动启用邮件发送功能"
            messagebox.showinfo("成功", msg)
            self.refresh_home_status()

    # ========== API密钥配置 ==========
    def create_api_keys_tab(self):
        frame = ttk.Frame(self.notebook, padding=20)
        self.notebook.add(frame, text="🔑 API密钥")

        ttk.Label(frame, text="学术API密钥配置", font=('Microsoft YaHei', 16, 'bold'), bootstyle=PRIMARY).pack(pady=(0, 5))
        ttk.Label(frame, text="配置API密钥可提升访问速率和配额，部分API必须配置密钥才能使用", bootstyle=SECONDARY).pack(pady=(0, 20))

        form_frame = ttk.Frame(frame)
        form_frame.pack(fill='x', pady=10)

        self.api_key_entries = {}
        api_key_config = [
            ('semantic_scholar', 'Semantic Scholar:', '可选 - 无Key限流100次/5分钟'),
            ('openalex', 'OpenAlex:', '可选 - 有Key可提升速率'),
            ('core', 'CORE:', '必须 - 无Key无法使用'),
        ]

        for i, (key_name, label, hint) in enumerate(api_key_config):
            row = ttk.Frame(form_frame)
            row.pack(fill='x', pady=10)

            ttk.Label(row, text=label, width=18, font=('Microsoft YaHei', 10)).pack(side='left')

            entry = ttk.Entry(row, width=50, show='•')
            entry.pack(side='left', padx=15)
            self.api_key_entries[key_name] = entry

            ttk.Label(row, text=hint, bootstyle=SECONDARY).pack(side='left', padx=5)

        self.show_api_keys = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame, text="显示密钥内容", variable=self.show_api_keys,
                        command=self._toggle_api_key_visibility, bootstyle="round-toggle").pack(pady=15)

        ttk.Button(frame, text="💾 保存API密钥配置",
                   command=self.save_api_keys_config, bootstyle=SUCCESS, width=20).pack(pady=10)

        help_frame = ttk.Labelframe(frame, text="🔗 获取API密钥指引", padding=15)
        help_frame.pack(fill='x', pady=20)

        links = [
            ("Semantic Scholar API Key", "https://www.semanticscholar.org/product/api#api-key"),
            ("OpenAlex API Key", "https://docs.openalex.org/how-to-use-the-api/get-an-api-key"),
            ("CORE API Key", "https://core.ac.uk/services/api"),
        ]

        for name, url in links:
            link_frame = ttk.Frame(help_frame)
            link_frame.pack(fill='x', pady=5)
            ttk.Label(link_frame, text=f"• {name}:", width=25).pack(side='left')
            link_label = ttk.Label(link_frame, text=url, bootstyle=PRIMARY, cursor='hand2')
            link_label.pack(side='left', padx=5)
            link_label.bind('<Button-1>', lambda e, u=url: webbrowser.open(u))

        self.api_key_status = ttk.Label(frame, text="", bootstyle=WARNING, font=('Microsoft YaHei', 10, 'bold'))
        self.api_key_status.pack(pady=15)

    def _toggle_api_key_visibility(self):
        show = self.show_api_keys.get()
        for entry in self.api_key_entries.values():
            entry.config(show='' if show else '•')

    def load_api_keys_config(self):
        if not hasattr(self, 'api_key_entries'):
            return
        keys = self.config_manager.get_api_keys()
        for key_name, entry in self.api_key_entries.items():
            entry.delete(0, tk.END)
            value = keys.get(key_name, '')
            if value and value.startswith('enc:'):
                value = _decode_pwd(value)
            entry.insert(0, value)

        self._check_api_key_status()

    def save_api_keys_config(self):
        api_keys = {}
        for key_name, entry in self.api_key_entries.items():
            value = entry.get().strip()
            if value and not value.startswith('enc:'):
                value = _encode_pwd(value)
            api_keys[key_name] = value

        if self.config_manager.update_api_keys(api_keys):
            messagebox.showinfo("成功", "API密钥配置已保存")
            self._check_api_key_status()
            self.refresh_home_status()

    def _check_api_key_status(self):
        if not hasattr(self, 'api_key_entries'):
            return
        missing = self.config_manager.check_missing_api_keys()
        if missing:
            required_missing = [m for m in missing if m['required']]
            optional_missing = [m for m in missing if not m['required']]
            msg_parts = []
            if required_missing:
                names = ', '.join(m['name'] for m in required_missing)
                msg_parts.append(f"⚠️ 必须配置: {names}")
            if optional_missing:
                names = ', '.join(m['name'] for m in optional_missing)
                msg_parts.append(f"💡 建议配置: {names}")
            self.api_key_status.config(text=' | '.join(msg_parts))
        else:
            self.api_key_status.config(text="✅ 所有API密钥已配置", bootstyle=SUCCESS)

    # ========== 课程表 ==========
    def create_schedule_tab(self):
        frame = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(frame, text="📅 课程表")

        info_frame = ttk.Labelframe(frame, text="学生信息", padding=10)
        info_frame.pack(fill='x', pady=5)

        info_row = ttk.Frame(info_frame)
        info_row.pack(fill='x', pady=5)

        ttk.Label(info_row, text="学校:").pack(side='left')
        self.school_var = tk.StringVar()
        self.school_combo = ttk.Combobox(info_row, textvariable=self.school_var, values=POPULAR_SCHOOLS, width=25)
        self.school_combo.pack(side='left', padx=10)
        self.school_combo.bind('<<ComboboxSelected>>', self._on_school_select)

        self.school_entry = ttk.Entry(info_row, width=25)
        self.school_entry.pack(side='left', padx=10)
        self.school_entry.pack_forget()

        ttk.Label(info_row, text="年级:").pack(side='left', padx=(20, 0))
        self.grade_var = tk.StringVar()
        ttk.Combobox(info_row, textvariable=self.grade_var, values=GRADES, width=15, state='readonly').pack(side='left', padx=10)

        ttk.Button(info_row, text="💾 保存信息", command=self._save_user_info, bootstyle=SUCCESS).pack(side='left', padx=20)

        vis_frame = ttk.Labelframe(frame, text="本周课程表", padding=10)
        vis_frame.pack(fill='both', expand=True, pady=10)

        self.canvas = tk.Canvas(vis_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(vis_frame, orient='vertical', command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)

        self.canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        self.canvas.bind('<Configure>', lambda e: self.canvas.configure(scrollregion=self.canvas.bbox('all')))

        self.timetable_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.timetable_frame, anchor='nw')

        self._build_weekly_timetable()

        # 当日课程详情
        select_frame = ttk.Frame(frame)
        select_frame.pack(fill='x', pady=5)

        ttk.Label(select_frame, text="选择日期:").pack(side='left')
        self.day_var = tk.StringVar()
        self.day_combo = ttk.Combobox(select_frame, textvariable=self.day_var, values=DAY_NAMES_CN, width=15, state='readonly')
        self.day_combo.current(0)
        self.day_combo.pack(side='left', padx=10)
        self.day_combo.bind('<<ComboboxSelected>>', self.on_day_change)

        list_frame = ttk.Labelframe(frame, text="当日课程详情", padding=10)
        list_frame.pack(fill='both', expand=True, pady=5)

        columns = ('name', 'time', 'location', 'teacher')
        self.schedule_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=6, bootstyle=INFO)

        self.schedule_tree.heading('name', text='课程名称')
        self.schedule_tree.heading('time', text='时间')
        self.schedule_tree.heading('location', text='地点')
        self.schedule_tree.heading('teacher', text='教师')

        self.schedule_tree.column('name', width=250)
        self.schedule_tree.column('time', width=150)
        self.schedule_tree.column('location', width=250)
        self.schedule_tree.column('teacher', width=150)

        self.schedule_tree.pack(fill='both', expand=True)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=10)

        ttk.Button(btn_frame, text="➕ 添加课程", command=self.add_course, bootstyle=SUCCESS).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🗑️ 删除课程", command=self.delete_course, bootstyle=DANGER).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📥 导入课表", command=self.import_schedule, bootstyle=PRIMARY).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🔄 刷新视图", command=self._refresh_timetable, bootstyle=INFO).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="⏰ 编辑时间段", command=self._edit_time_slots, bootstyle=WARNING).pack(side='left', padx=5)

        self._load_user_info()

    def _on_school_select(self, event):
        if self.school_var.get() == "其他（手动输入）":
            self.school_combo.pack_forget()
            self.school_entry.pack(side='left', padx=10)
            self.school_entry.focus_set()
        else:
            self.school_entry.pack_forget()
            self.school_combo.pack(side='left', padx=10)

    def _load_user_info(self):
        user_info = self.config_manager.get_user_info()
        school = user_info.get('university', '')
        grade = user_info.get('grade', '')
        if school in POPULAR_SCHOOLS:
            self.school_var.set(school)
        elif school:
            self.school_var.set("其他（手动输入）")
            self.school_entry.pack(side='left', padx=10)
            self.school_entry.insert(0, school)
        self.grade_var.set(grade)

    def _save_user_info(self):
        school = self.school_entry.get().strip() if self.school_var.get() == "其他（手动输入）" else self.school_var.get()
        grade = self.grade_var.get()
        self.config_manager.update_user_info(university=school, grade=grade)
        messagebox.showinfo("成功", "学生信息已保存")

    def _build_weekly_timetable(self):
        for w in self.timetable_frame.winfo_children():
            w.destroy()

        time_slot_configs = self.config_manager.get_time_slots()
        time_slots = [f"{s['label']}\n{s['start']}-{s['end']}" for s in time_slot_configs]
        days = DAY_NAMES_CN
        colors = ['#e3f2fd', '#f3e5f5', '#e8f5e9', '#fff3e0', '#fce4ec', '#e0f7fa', '#f9fbe7']

        cell_w = 110
        cell_h = 75
        header_h = 40
        time_w = 100

        canvas_w = time_w + len(days) * cell_w + 10
        canvas_h = header_h + len(time_slots) * cell_h + 10
        self.canvas.configure(width=canvas_w, height=min(canvas_h, 400))

        header_font = ('Microsoft YaHei', 10, 'bold')
        cell_font = ('Microsoft YaHei', 9)

        # Draw onto Canvas directly inside the frame context
        canvas = tk.Canvas(self.timetable_frame, width=canvas_w, height=canvas_h, bg='white', highlightthickness=0)
        canvas.pack()

        canvas.create_text(time_w // 2, header_h // 2, text="时间/日期", font=header_font, anchor='center')

        for j, day in enumerate(days):
            x = time_w + j * cell_w + cell_w // 2
            canvas.create_text(x, header_h // 2, text=day, font=header_font, anchor='center')

        for i, slot in enumerate(time_slots):
            y = header_h + i * cell_h + cell_h // 2
            canvas.create_text(time_w // 2, y, text=slot, font=cell_font, anchor='center')

        schedule = self.config_manager.get_week_schedule()

        for j, day_cn in enumerate(days):
            day_en = DAY_MAP.get(day_cn, 'Monday')
            courses = schedule.get(day_en, [])

            for course in courses:
                time_str = course.get('time', '')
                slot_idx = self._time_to_slot(time_str)
                if slot_idx < 0:
                    slot_idx = i

                x1 = time_w + j * cell_w + 2
                y1 = header_h + slot_idx * cell_h + 2
                x2 = x1 + cell_w - 4
                y2 = y1 + cell_h - 4

                color = colors[j % len(colors)]
                canvas.create_rectangle(x1, y1, x2, y2, fill=color, outline='#e0e0e0', width=1)

                name = course.get('name', '')[:8]
                loc = course.get('location', '')[:6]
                canvas.create_text(x1 + (cell_w - 4) // 2, y1 + 20, text=name, font=cell_font, anchor='center')
                canvas.create_text(x1 + (cell_w - 4) // 2, y1 + 45, text=loc, font=('Microsoft YaHei', 8), fill='#666', anchor='center')

        for i in range(len(time_slots) + 1):
            y = header_h + i * cell_h
            canvas.create_line(0, y, canvas_w, y, fill='#eee')
        for j in range(len(days) + 1):
            x = time_w + j * cell_w
            canvas.create_line(x, 0, x, canvas_h, fill='#eee')

    def _time_to_slot(self, time_str):
        if not time_str:
            return 0
        time_slot_configs = self.config_manager.get_time_slots()
        time_lower = time_str.lower()
        for idx, slot in enumerate(time_slot_configs):
            label = slot['label'].lower()
            start = slot['start'].lower()
            end = slot['end'].lower()
            if label in time_lower or start in time_lower or end in time_lower:
                return idx
        return 0

    def _refresh_timetable(self):
        self._build_weekly_timetable()
        self.on_day_change(None)

    def load_schedule(self):
        self.on_day_change(None)

    def on_day_change(self, event):
        for item in self.schedule_tree.get_children():
            self.schedule_tree.delete(item)

        day_name = self.day_var.get()
        day = DAY_MAP.get(day_name, 'Monday')

        schedule = self.config_manager.get_week_schedule()
        courses = schedule.get(day, [])

        for course in courses:
            self.schedule_tree.insert('', 'end', values=(
                course.get('name', ''),
                course.get('time', ''),
                course.get('location', ''),
                course.get('teacher', '')
            ))

    def add_course(self):
        day_name = self.day_var.get()
        day = DAY_MAP.get(day_name, 'Monday')

        dialog = CourseDialog(self.root, "添加课程")
        if dialog.result:
            self.config_manager.add_course(day, **dialog.result)
            self.on_day_change(None)
            messagebox.showinfo("成功", "课程已添加")

    def delete_course(self):
        selected = self.schedule_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的课程")
            return

        if messagebox.askyesno("确认", "确定要删除此课程吗？"):
            day_name = self.day_var.get()
            day = DAY_MAP.get(day_name, 'Monday')
            index = self.schedule_tree.index(selected[0])
            self.config_manager.delete_course(day, index)
            self.on_day_change(None)

    def import_schedule(self):
        dialog = ScheduleImportDialog(self.root, self.config_manager)
        if dialog.imported:
            self.on_day_change(None)
            self.refresh_home_status()

    def _edit_time_slots(self):
        dialog = TimeSlotDialog(self.root, self.config_manager)
        if dialog.saved:
            self._build_weekly_timetable()
            self.on_day_change(None)

    # ========== 关于 ==========
    def create_about_tab(self):
        frame = ttk.Frame(self.notebook, padding=40)
        self.notebook.add(frame, text="ℹ️ 关于")

        ttk.Label(frame, text="📚 Daily Automation", font=('Microsoft YaHei', 26, 'bold'), bootstyle=PRIMARY).pack(pady=(20, 10))
        ttk.Label(frame, text="学术自动化助手 v1.0", font=('Microsoft YaHei', 14), bootstyle=INFO).pack()
        ttk.Label(frame, text="让信息找上门，省下时间做重要的事", font=('Microsoft YaHei', 11), bootstyle=SECONDARY).pack(pady=15)

        ttk.Separator(frame, orient='horizontal').pack(fill='x', padx=100, pady=30)

        features = ttk.Labelframe(frame, text="功能特性", padding=20)
        features.pack(fill='x', padx=100, pady=10)

        feature_text = """
✅ 多源学术信息爬取（arXiv、Semantic Scholar等）
✅ 智能关键词筛选
✅ 每日简报生成
✅ 邮件自动发送
✅ 日程提醒
✅ 课程表管理
        """
        ttk.Label(features, text=feature_text, justify='left', font=('Microsoft YaHei', 11)).pack()

        ttk.Button(frame, text="🔄 重置为默认配置", command=self.reset_config, bootstyle=(DANGER, OUTLINE)).pack(pady=40)

    def reset_config(self):
        if messagebox.askyesno("确认", "确定要重置为默认配置吗？所有自定义设置将丢失。"):
            self.config_manager.reset_to_default()
            self.load_all_config()
            messagebox.showinfo("成功", "已重置为默认配置")

    # ========== 通用方法 ==========
    def load_all_config(self):
        self.load_sources()
        self.load_keywords()
        self.load_reminders()
        self.load_email_config()
        self.load_api_keys_config()
        self.load_schedule()
        self.refresh_home_status()

    def _run_subprocess_task(self, task_mode: str, status_msg: str, success_msg: str, is_async: bool = False):
        dialog = _TaskProgressDialog(self.root, self.app_dir, self, task_mode, status_msg, success_msg, is_async=is_async)
        dialog.run()

    def run_once(self):
        self._run_subprocess_task("all", "正在运行...", "运行完成")

    def run_once_async(self):
        self._run_subprocess_task("all", "正在异步运行...", "异步运行完成", is_async=True)

    def run_crawl(self):
        self._run_subprocess_task("crawl", "正在生成学术简报...", "学术简报生成完成")

    def run_remind(self):
        self._run_subprocess_task("remind", "正在检查日程提醒...", "日程提醒检查完成")

    def open_data_dir(self):
        data_dir = self.app_dir / "data"
        data_dir.mkdir(exist_ok=True)
        webbrowser.open(str(data_dir))

    def export_config(self):
        data = self.config_manager.export_all_config()
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")],
            initialfile=f"daily_automation_config_{datetime.now().strftime('%Y%m%d')}.json"
        )
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("成功", f"配置已导出到: {filename}")

    def import_config(self):
        filename = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")]
        )
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if self.config_manager.import_all_config(data):
                    self.load_all_config()
                    messagebox.showinfo("成功", "配置已导入")
            except Exception as e:
                self._safe_showerror("错误", f"导入失败: {e}")


# ========== 首次启动向导 ==========

class SetupWizard:
    def __init__(self, parent, config_manager):
        self.config_manager = config_manager
        self.completed = False
        self.current_step = 0
        self.total_steps = 4
        self.step_titles = ["基本信息", "课表导入", "数据源选择", "兴趣方向"]
        self.selected_interests = []

        self.wizard = ttk.Toplevel(parent, title="欢迎使用 Daily Automation")
        self.wizard.geometry("800x650")
        self.wizard.resizable(True, True)
        self.wizard.transient(parent)
        self.wizard.grab_set()
        self.wizard.protocol("WM_DELETE_WINDOW", self._on_skip)

        header = ttk.Frame(self.wizard)
        header.pack(fill='x', padx=30, pady=(20, 10))
        ttk.Label(header, text="📚 Daily Automation 初始设置",
                  font=('Microsoft YaHei', 20, 'bold'), bootstyle=PRIMARY).pack()
        ttk.Label(header, text="快速完成配置，3步即可开始使用", bootstyle=SECONDARY).pack(pady=5)

        self.step_indicator = ttk.Frame(self.wizard)
        self.step_indicator.pack(fill='x', padx=30, pady=10)
        self.step_labels = []
        for i, title in enumerate(self.step_titles):
            lbl = ttk.Label(self.step_indicator, text=f"  {i+1}. {title}  ", font=('Microsoft YaHei', 11))
            lbl.pack(side='left', padx=10)
            self.step_labels.append(lbl)

        ttk.Separator(self.wizard, orient='horizontal').pack(fill='x', padx=30, pady=10)

        self.content_frame = ttk.Frame(self.wizard, padding=10)
        self.content_frame.pack(fill='both', expand=True, padx=30, pady=5)

        btn_frame = ttk.Frame(self.wizard, padding=20)
        btn_frame.pack(fill='x')

        self.skip_btn = ttk.Button(btn_frame, text="跳过全部", command=self._on_skip, bootstyle=SECONDARY)
        self.skip_btn.pack(side='left')

        self.back_btn = ttk.Button(btn_frame, text="◀ 上一步", command=self._prev_step, state='disabled', bootstyle=INFO)
        self.back_btn.pack(side='right', padx=10)

        self.next_btn = ttk.Button(btn_frame, text="下一步 ▶", command=self._next_step, bootstyle=PRIMARY)
        self.next_btn.pack(side='right', padx=10)

        self._build_step(0)
        self._update_step_indicator()
        self.wizard.wait_window()

    def _build_step(self, step):
        for w in self.content_frame.winfo_children():
            w.destroy()

        if step == 0:
            self._build_step_basic()
        elif step == 1:
            self._build_step_schedule()
        elif step == 2:
            self._build_step_sources()
        elif step == 3:
            self._build_step_interests()

    def _build_step_basic(self):
        f = self.content_frame

        ttk.Label(f, text="📝 基本信息设置", font=('Microsoft YaHei', 16, 'bold'), bootstyle=PRIMARY).pack(anchor='w', pady=(5, 20))

        row1 = ttk.Frame(f)
        row1.pack(fill='x', pady=10)
        ttk.Label(row1, text="学校:", width=10, font=('Microsoft YaHei', 11)).pack(side='left')
        self.school_var = tk.StringVar()
        self.school_combo = ttk.Combobox(row1, textvariable=self.school_var, values=POPULAR_SCHOOLS, width=35)
        self.school_combo.pack(side='left', padx=10)
        self.school_combo.bind('<<ComboboxSelected>>', self._on_school_select)

        self.school_entry = ttk.Entry(row1, width=35)
        self.school_entry.pack(side='left', padx=10)
        self.school_entry.pack_forget()

        row2 = ttk.Frame(f)
        row2.pack(fill='x', pady=10)
        ttk.Label(row2, text="年级:", width=10, font=('Microsoft YaHei', 11)).pack(side='left')
        self.grade_var = tk.StringVar()
        ttk.Combobox(row2, textvariable=self.grade_var, values=GRADES, width=33, state='readonly').pack(side='left', padx=10)

        row3 = ttk.Frame(f)
        row3.pack(fill='x', pady=10)
        ttk.Label(row3, text="学期:", width=10, font=('Microsoft YaHei', 11)).pack(side='left')
        self.semester_var = tk.StringVar()
        now = datetime.now()
        year = now.year
        sem = "春夏" if now.month >= 2 and now.month <= 7 else "秋冬"
        self.semester_var.set(f"{year-1}-{year}学年{sem}学期")
        ttk.Entry(row3, textvariable=self.semester_var, width=35).pack(side='left', padx=10)

        ttk.Label(f, text="💡 这些信息仅用于个性化你的每日计划，不会上传到任何服务器",
                  bootstyle=WARNING, font=('Microsoft YaHei', 10)).pack(anchor='w', pady=(30, 0))

    def _on_school_select(self, event):
        if self.school_var.get() == "其他（手动输入）":
            self.school_combo.pack_forget()
            self.school_entry.pack(side='left', padx=10)
            self.school_entry.focus_set()
        else:
            self.school_entry.pack_forget()
            self.school_combo.pack(side='left', padx=10)

    def _build_step_schedule(self):
        f = self.content_frame

        ttk.Label(f, text="📅 课表导入", font=('Microsoft YaHei', 16, 'bold'), bootstyle=PRIMARY).pack(anchor='w', pady=(5, 10))
        ttk.Label(f, text="从教务系统复制课表粘贴到下方，或选择CSV/Excel文件导入", bootstyle=SECONDARY).pack(anchor='w', pady=(0, 15))

        method_frame = ttk.Frame(f)
        method_frame.pack(fill='x', pady=10)

        self.schedule_method = tk.StringVar(value='text')
        ttk.Radiobutton(method_frame, text="📝 粘贴文字", variable=self.schedule_method,
                        value='text', command=self._toggle_schedule_method).pack(side='left', padx=10)
        ttk.Radiobutton(method_frame, text="📂 导入文件", variable=self.schedule_method,
                        value='file', command=self._toggle_schedule_method).pack(side='left', padx=10)

        self.text_frame = ttk.Frame(f)
        self.text_frame.pack(fill='both', expand=True, pady=10)

        ttk.Label(self.text_frame, text="粘贴课表内容（支持从教务系统直接复制）:", font=('Microsoft YaHei', 10)).pack(anchor='w', pady=5)
        self.schedule_text = ScrolledText(self.text_frame, height=8, width=65)
        self.schedule_text.pack(fill='both', expand=True, pady=5)
        ttk.Label(self.text_frame, text="支持格式：Tab分隔 / 逗号分隔 / 竖线分隔，含列名或不含均可",
                  bootstyle=SECONDARY).pack(anchor='w')

        self.file_frame = ttk.Frame(f)

        self.file_path_var = tk.StringVar()
        ttk.Entry(self.file_frame, textvariable=self.file_path_var, width=50).pack(side='left', padx=10)
        ttk.Button(self.file_frame, text="浏览...", command=self._browse_schedule_file, bootstyle=INFO).pack(side='left', padx=5)
        ttk.Label(self.file_frame, text="支持 CSV / Excel (.xlsx) 文件", bootstyle=SECONDARY).pack(side='left', padx=10)

        self.preview_frame = ttk.Labelframe(f, text="解析预览", padding=10)
        self.preview_frame.pack(fill='both', expand=True, pady=15)

        self.preview_tree = ttk.Treeview(self.preview_frame, columns=('day', 'name', 'time', 'location', 'teacher'),
                                         show='headings', height=5, bootstyle=INFO)
        for col, heading, width in [('day', '星期', 60), ('name', '课程名', 120), ('time', '时间', 100),
                                     ('location', '地点', 100), ('teacher', '教师', 80)]:
            self.preview_tree.heading(col, text=heading)
            self.preview_tree.column(col, width=width)
        self.preview_tree.pack(fill='both', expand=True)

        btn_row = ttk.Frame(f)
        btn_row.pack(fill='x', pady=5)
        ttk.Button(btn_row, text="🔍 解析预览", command=self._preview_schedule, bootstyle=PRIMARY).pack(side='left', padx=5)

    def _toggle_schedule_method(self):
        if self.schedule_method.get() == 'text':
            self.file_frame.pack_forget()
            self.text_frame.pack(fill='both', expand=True, pady=10)
        else:
            self.text_frame.pack_forget()
            self.file_frame.pack(fill='x', pady=10)

    def _browse_schedule_file(self):
        filetypes = [("表格文件", "*.csv *.xlsx *.xls"), ("CSV文件", "*.csv"), ("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            self.file_path_var.set(filename)

    def _preview_schedule(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)

        courses = self._get_parsed_courses()
        day_cn_map = {v: k for k, v in DAY_MAP.items()}
        for c in courses:
            day_cn = day_cn_map.get(c['day'], c['day'])
            self.preview_tree.insert('', 'end', values=(day_cn, c['name'], c['time'], c['location'], c['teacher']))

        if not courses:
            messagebox.showinfo("提示", "未能解析出课程信息，请检查输入格式")

    def _get_parsed_courses(self):
        if self.schedule_method.get() == 'text':
            text = self.schedule_text.get(1.0, tk.END).strip()
            if not text:
                return []
            return parse_course_table(text)
        else:
            filepath = self.file_path_var.get().strip()
            if not filepath:
                return []
            return self._parse_file(filepath)

    def _parse_file(self, filepath):
        path = Path(filepath)
        if not path.exists():
            return []

        try:
            if path.suffix.lower() == '.csv':
                return self._parse_csv_file(path)
            elif path.suffix.lower() in ('.xlsx', '.xls'):
                return self._parse_excel_file(path)
            else:
                with open(path, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
                return parse_course_table(content)
        except Exception as e:
            self._safe_showerror("解析错误", f"文件解析失败: {e}")
            return []

    def _parse_csv_file(self, path):
        courses = []
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            return courses

        text = '\n'.join(['\t'.join(row) for row in rows])
        return parse_course_table(text)

    def _parse_excel_file(self, path):
        try:
            import openpyxl
        except ImportError:
            messagebox.showwarning("提示", "Excel文件需要安装 openpyxl 库。\n请运行: pip install openpyxl\n或改用CSV格式导入。")
            return []

        courses = []
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else '' for c in row]
            rows.append(cells)
        wb.close()

        if not rows:
            return courses

        text = '\n'.join(['\t'.join(row) for row in rows])
        return parse_course_table(text)

    def _build_step_sources(self):
        f = self.content_frame

        ttk.Label(f, text="📡 数据源选择", font=('Microsoft YaHei', 16, 'bold'), bootstyle=PRIMARY).pack(anchor='w', pady=(5, 10))
        ttk.Label(f, text="选择你关注的学术数据源（取消勾选可减少运行时间）", bootstyle=SECONDARY).pack(anchor='w', pady=(0, 15))

        canvas_frame = ttk.Frame(f)
        canvas_frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(canvas_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient='vertical', command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all('<MouseWheel>', _on_mousewheel)

        self.source_vars = {}
        current_category = None

        for source in DEFAULT_NEWS_SOURCES:
            cat_key = source.get('category', 'general_academic')
            cat_name = SOURCE_CATEGORIES.get(cat_key, cat_key)

            if cat_key != current_category:
                current_category = cat_key
                cat_label = ttk.Label(scroll_frame, text=f"  {cat_name}",
                                      font=('Microsoft YaHei', 11, 'bold'), bootstyle=INFO)
                cat_label.pack(anchor='w', padx=5, pady=(15, 5))
                ttk.Separator(scroll_frame, orient='horizontal').pack(fill='x', padx=5, pady=5)

            row_frame = ttk.Frame(scroll_frame)
            row_frame.pack(fill='x', padx=15, pady=5)

            var = tk.BooleanVar(value=source.get('enabled', True))
            self.source_vars[source['name']] = var

            cb = ttk.Checkbutton(row_frame, text=source['name'], variable=var, bootstyle="round-toggle")
            cb.pack(side='left')

            desc = source.get('desc', '')
            if desc:
                desc_label = ttk.Label(row_frame, text=f"  - {desc}", bootstyle=SECONDARY, font=('Microsoft YaHei', 9))
                desc_label.pack(side='left', padx=(10, 0))

            type_tag = source.get('type', 'web').upper()
            type_label = ttk.Label(row_frame, text=f"[{type_tag}]", bootstyle=PRIMARY, font=('Consolas', 9))
            type_label.pack(side='right', padx=10)

        btn_frame = ttk.Frame(f)
        btn_frame.pack(fill='x', pady=15)
        ttk.Button(btn_frame, text="全选", command=lambda: self._set_all_sources(True), bootstyle=SUCCESS).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="全不选", command=lambda: self._set_all_sources(False), bootstyle=DANGER).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="推荐配置", command=self._set_recommended_sources, bootstyle=PRIMARY).pack(side='left', padx=5)

    def _set_all_sources(self, value):
        if hasattr(self, 'source_vars'):
            for var in self.source_vars.values():
                var.set(value)

    def _set_recommended_sources(self):
        if hasattr(self, 'source_vars'):
            for name, var in self.source_vars.items():
                default = next((s['enabled'] for s in DEFAULT_NEWS_SOURCES if s['name'] == name), True)
                var.set(default)

    def _build_step_interests(self):
        f = self.content_frame

        ttk.Label(f, text="🎯 兴趣方向", font=('Microsoft YaHei', 16, 'bold'), bootstyle=PRIMARY).pack(anchor='w', pady=(5, 10))
        ttk.Label(f, text="选择你感兴趣的领域，将自动填充相关学术关键词（可多选）", bootstyle=SECONDARY).pack(anchor='w', pady=(0, 15))

        tags_frame = ttk.Frame(f)
        tags_frame.pack(fill='x', pady=10)

        self.interest_vars = {}
        for i, (name, _, _) in enumerate(INTEREST_AREAS):
            row = i // 3
            col = i % 3
            var = tk.BooleanVar(value=False)
            self.interest_vars[name] = var
            cb = ttk.Checkbutton(tags_frame, text=name, variable=var, bootstyle="primary")
            cb.grid(row=row, column=col, sticky='w', padx=15, pady=8)

        ttk.Separator(f, orient='horizontal').pack(fill='x', pady=20)

        custom_frame = ttk.Labelframe(f, text="自定义关键词（可选）", padding=15)
        custom_frame.pack(fill='both', expand=True, pady=10)

        dual = ttk.Frame(custom_frame)
        dual.pack(fill='both', expand=True, padx=5, pady=5)

        ttk.Label(dual, text="英文关键词（每行一个）:").pack(anchor='w', pady=(0, 5))
        self.custom_en = ScrolledText(dual, height=4, width=30)
        self.custom_en.pack(fill='both', expand=True, side='left', padx=(0, 10))

        ttk.Label(dual, text="中文关键词（每行一个）:").pack(anchor='w', pady=(0, 5))
        self.custom_cn = ScrolledText(dual, height=4, width=30)
        self.custom_cn.pack(fill='both', expand=True, side='left')

    def _update_step_indicator(self):
        for i, lbl in enumerate(self.step_labels):
            if i < self.current_step:
                lbl.config(bootstyle=SUCCESS)
            elif i == self.current_step:
                lbl.config(bootstyle=PRIMARY, font=('Microsoft YaHei', 11, 'bold'))
            else:
                lbl.config(bootstyle=SECONDARY)

    def _next_step(self):
        if self.current_step == 1:
            courses = self._get_parsed_courses()
            if courses:
                schedule = self.config_manager.get_schedule()
                if 'week_schedule' not in schedule:
                    schedule['week_schedule'] = {}
                for day in DAY_NAMES_EN:
                    if day not in schedule['week_schedule']:
                        schedule['week_schedule'][day] = []

                for c in courses:
                    day = c['day']
                    if day not in schedule['week_schedule']:
                        schedule['week_schedule'][day] = []
                    schedule['week_schedule'][day].append({
                        'name': c['name'],
                        'time': c['time'],
                        'location': c['location'],
                        'teacher': c['teacher'],
                        'note': c.get('note', '')
                    })
                self.config_manager.save_schedule(schedule)

        if self.current_step == 2:
            self._save_source_selection()

        if self.current_step < self.total_steps - 1:
            self.current_step += 1
            self._build_step(self.current_step)
            self._update_step_indicator()
            self.back_btn.config(state='normal')
            if self.current_step == self.total_steps - 1:
                self.next_btn.config(text="完成", bootstyle=SUCCESS)
        else:
            self._save_basic_info()
            self._save_interests()
            self.completed = True
            self.wizard.destroy()

    def _prev_step(self):
        if self.current_step > 0:
            self.current_step -= 1
            self._build_step(self.current_step)
            self._update_step_indicator()
            self.back_btn.config(state='disabled' if self.current_step == 0 else 'normal')
            self.next_btn.config(text="下一步 ▶", bootstyle=PRIMARY)

    def _on_skip(self):
        self._save_basic_info()
        self.completed = True
        self.wizard.destroy()

    def _save_basic_info(self):
        if not hasattr(self, 'semester_var'):
            return

        semester = self.semester_var.get().strip() if hasattr(self, 'semester_var') else ""
        if semester:
            schedule = self.config_manager.get_schedule()
            schedule['semester'] = semester
            self.config_manager.save_schedule(schedule)

        config = self.config_manager.get_config()
        if not config.get('reminders'):
            config['reminders'] = [
                {"time": "09:00", "title": "学术早报", "description": "今日最新学术资讯已整理完毕，请查阅。"},
                {"time": "22:00", "title": "晚间复盘", "description": "今日工作告一段落，回顾一下收获。"}
            ]
            self.config_manager.save_config(config)

    def _save_source_selection(self):
        if not hasattr(self, 'source_vars'):
            return

        config = self.config_manager.get_config()
        sources = config.get('news_sources', [])

        for source in sources:
            if source['name'] in self.source_vars:
                source['enabled'] = self.source_vars[source['name']].get()

        config['news_sources'] = sources
        self.config_manager.save_config(config)

    def _save_interests(self):
        if not hasattr(self, 'interest_vars'):
            return

        en_keywords = []
        cn_keywords = []

        for name, en_list, cn_list in INTEREST_AREAS:
            if self.interest_vars.get(name, tk.BooleanVar()).get():
                en_keywords.extend(en_list)
                cn_keywords.extend(cn_list)

        if hasattr(self, 'custom_en'):
            custom_en = [k.strip() for k in self.custom_en.get(1.0, tk.END).split('\n') if k.strip()]
            custom_cn = [k.strip() for k in self.custom_cn.get(1.0, tk.END).split('\n') if k.strip()]
            en_keywords.extend(custom_en)
            cn_keywords.extend(custom_cn)

        en_keywords = list(dict.fromkeys(en_keywords))
        cn_keywords = list(dict.fromkeys(cn_keywords))

        if en_keywords or cn_keywords:
            config = self.config_manager.get_config()
            existing_en = config.get('keywords', [])
            existing_cn = config.get('keywords_cn', [])
            merged_en = list(dict.fromkeys(existing_en + en_keywords))
            merged_cn = list(dict.fromkeys(existing_cn + cn_keywords))
            self.config_manager.update_keywords(merged_en, merged_cn)


# ========== 课表导入对话框 ==========

class ScheduleImportDialog:
    def __init__(self, parent, config_manager):
        self.config_manager = config_manager
        self.imported = False

        self.dialog = ttk.Toplevel(parent, title="📥 导入课表")
        self.dialog.geometry("750x600")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        ttk.Label(self.dialog, text="导入课表", font=('Microsoft YaHei', 16, 'bold'), bootstyle=PRIMARY).pack(pady=(20, 5))
        ttk.Label(self.dialog, text="从教务系统复制课表粘贴到下方，或选择CSV/Excel文件导入", bootstyle=SECONDARY).pack(pady=(0, 15))

        method_frame = ttk.Frame(self.dialog)
        method_frame.pack(fill='x', padx=25, pady=5)

        self.method = tk.StringVar(value='text')
        ttk.Radiobutton(method_frame, text="📝 粘贴文字", variable=self.method,
                        value='text', command=self._toggle_method).pack(side='left', padx=10)
        ttk.Radiobutton(method_frame, text="📂 导入文件", variable=self.method,
                        value='file', command=self._toggle_method).pack(side='left', padx=10)

        self.text_frame = ttk.Frame(self.dialog)
        self.text_frame.pack(fill='both', expand=True, padx=25, pady=10)

        ttk.Label(self.text_frame, text="粘贴课表内容（支持从教务系统直接复制）:").pack(anchor='w', pady=(0, 5))
        self.text_area = ScrolledText(self.text_frame, height=6, width=70)
        self.text_area.pack(fill='both', expand=True, pady=3)
        ttk.Label(self.text_frame, text="支持格式：Tab分隔 / 逗号分隔 / 竖线分隔", bootstyle=SECONDARY).pack(anchor='w')

        self.file_frame = ttk.Frame(self.dialog)
        self.file_path = tk.StringVar()
        ttk.Entry(self.file_frame, textvariable=self.file_path, width=55).pack(side='left', padx=5)
        ttk.Button(self.file_frame, text="浏览...", command=self._browse_file, bootstyle=INFO).pack(side='left', padx=5)
        ttk.Label(self.file_frame, text="支持 CSV / Excel (.xlsx)", bootstyle=SECONDARY).pack(side='left', padx=5)

        preview_lf = ttk.Labelframe(self.dialog, text="解析预览", padding=10)
        preview_lf.pack(fill='both', expand=True, padx=25, pady=15)

        self.preview_tree = ttk.Treeview(preview_lf, columns=('day', 'name', 'time', 'location', 'teacher'),
                                         show='headings', height=6, bootstyle=INFO)
        for col, heading, width in [('day', '星期', 60), ('name', '课程名', 140), ('time', '时间', 100),
                                     ('location', '地点', 120), ('teacher', '教师', 80)]:
            self.preview_tree.heading(col, text=heading)
            self.preview_tree.column(col, width=width)
        self.preview_tree.pack(fill='both', expand=True)

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(fill='x', padx=25, pady=20)

        ttk.Button(btn_frame, text="🔍 解析预览", command=self._preview, bootstyle=INFO).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✅ 确认导入", command=self._confirm, bootstyle=SUCCESS).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="取消", command=self.dialog.destroy, bootstyle=SECONDARY).pack(side='right', padx=5)

        self.parsed_courses = []
        self.dialog.wait_window()

    def _toggle_method(self):
        if self.method.get() == 'text':
            self.file_frame.pack_forget()
            self.text_frame.pack(fill='both', expand=True, padx=25, pady=10)
        else:
            self.text_frame.pack_forget()
            self.file_frame.pack(fill='x', padx=25, pady=10)

    def _browse_file(self):
        filetypes = [("表格文件", "*.csv *.xlsx *.xls"), ("CSV文件", "*.csv"), ("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            self.file_path.set(filename)

    def _preview(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)

        self.parsed_courses = self._parse()
        day_cn_map = {v: k for k, v in DAY_MAP.items()}
        for c in self.parsed_courses:
            day_cn = day_cn_map.get(c['day'], c['day'])
            self.preview_tree.insert('', 'end', values=(day_cn, c['name'], c['time'], c['location'], c['teacher']))

        if not self.parsed_courses:
            messagebox.showinfo("提示", "未能解析出课程信息，请检查输入格式")

    def _parse(self):
        if self.method.get() == 'text':
            text = self.text_area.get(1.0, tk.END).strip()
            if not text:
                return []
            return parse_course_table(text)
        else:
            filepath = self.file_path.get().strip()
            if not filepath:
                return []
            path = Path(filepath)
            if not path.exists():
                return []
            try:
                if path.suffix.lower() == '.csv':
                    return self._parse_csv(path)
                elif path.suffix.lower() in ('.xlsx', '.xls'):
                    return self._parse_excel(path)
                else:
                    with open(path, 'r', encoding='utf-8', errors='replace') as f:
                        return parse_course_table(f.read())
            except Exception as e:
                self._safe_showerror("解析错误", f"文件解析失败: {e}")
                return []

    def _parse_csv(self, path):
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            rows = list(csv.reader(f))
        if not rows:
            return []
        text = '\n'.join(['\t'.join(row) for row in rows])
        return parse_course_table(text)

    def _parse_excel(self, path):
        try:
            import openpyxl
        except ImportError:
            messagebox.showwarning("提示", "Excel文件需要安装 openpyxl 库。\n请运行: pip install openpyxl\n或改用CSV格式导入。")
            return []
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else '' for c in row]
            rows.append(cells)
        wb.close()
        if not rows:
            return []
        text = '\n'.join(['\t'.join(row) for row in rows])
        return parse_course_table(text)

    def _confirm(self):
        if not self.parsed_courses:
            self._preview()
            if not self.parsed_courses:
                return

        schedule = self.config_manager.get_schedule()
        if 'week_schedule' not in schedule:
            schedule['week_schedule'] = {}
        for day in DAY_NAMES_EN:
            if day not in schedule['week_schedule']:
                schedule['week_schedule'][day] = []

        for c in self.parsed_courses:
            day = c['day']
            if day not in schedule['week_schedule']:
                schedule['week_schedule'][day] = []
            schedule['week_schedule'][day].append({
                'name': c['name'],
                'time': c['time'],
                'location': c['location'],
                'teacher': c['teacher'],
                'note': c.get('note', '')
            })

        self.config_manager.save_schedule(schedule)
        self.imported = True
        messagebox.showinfo("成功", f"已导入 {len(self.parsed_courses)} 门课程")
        self.dialog.destroy()


# ========== 对话框类（带输入验证） ==========

class SourceDialog:
    def __init__(self, parent, title, source=None):
        self.result = None

        self.dialog = ttk.Toplevel(parent, title=title)
        self.dialog.geometry("550x280")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog, padding=20)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text="名称:", font=('Microsoft YaHei', 10)).grid(row=0, column=0, sticky='e', pady=10, padx=10)
        self.name_entry = ttk.Entry(frame, width=45)
        self.name_entry.grid(row=0, column=1, pady=10, sticky='w')

        ttk.Label(frame, text="URL:", font=('Microsoft YaHei', 10)).grid(row=1, column=0, sticky='e', pady=10, padx=10)
        self.url_entry = ttk.Entry(frame, width=45)
        self.url_entry.grid(row=1, column=1, pady=10, sticky='w')

        ttk.Label(frame, text="类型:", font=('Microsoft YaHei', 10)).grid(row=2, column=0, sticky='e', pady=10, padx=10)
        self.type_var = tk.StringVar(value='rss')
        ttk.Combobox(frame, textvariable=self.type_var, values=['rss', 'web', 'json'], width=42, state='readonly').grid(row=2, column=1, pady=10, sticky='w')

        self.error_label = ttk.Label(frame, text="", bootstyle=DANGER, font=('Microsoft YaHei', 9))
        self.error_label.grid(row=3, column=0, columnspan=2, sticky='w', pady=(10, 0), padx=10)

        if source:
            self.name_entry.insert(0, source.get('name', ''))
            self.url_entry.insert(0, source.get('url', ''))
            self.type_var.set(source.get('type', 'rss'))

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="确定", command=self.ok, bootstyle=PRIMARY, width=12).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel, bootstyle=SECONDARY, width=12).pack(side='left', padx=10)

        self.dialog.wait_window()

    def ok(self):
        name = self.name_entry.get().strip()
        url = self.url_entry.get().strip()

        if not name:
            self.error_label.config(text="⚠ 请输入名称")
            return
        if not url:
            self.error_label.config(text="⚠ 请输入URL")
            return
        if not url.startswith(('http://', 'https://')):
            self.error_label.config(text="⚠ URL必须以 http:// 或 https:// 开头")
            return

        self.result = {
            'name': name,
            'url': url,
            'source_type': self.type_var.get(),
            'enabled': True
        }
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


class ReminderDialog:
    def __init__(self, parent, title, reminder=None):
        self.result = None

        self.dialog = ttk.Toplevel(parent, title=title)
        self.dialog.geometry("500x300")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog, padding=20)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text="时间 (HH:MM):", font=('Microsoft YaHei', 10)).grid(row=0, column=0, sticky='e', pady=10, padx=10)
        self.time_entry = ttk.Entry(frame, width=35)
        self.time_entry.grid(row=0, column=1, pady=10)

        ttk.Label(frame, text="标题:", font=('Microsoft YaHei', 10)).grid(row=1, column=0, sticky='e', pady=10, padx=10)
        self.title_entry = ttk.Entry(frame, width=35)
        self.title_entry.grid(row=1, column=1, pady=10)

        ttk.Label(frame, text="描述:", font=('Microsoft YaHei', 10)).grid(row=2, column=0, sticky='e', pady=10, padx=10)
        self.desc_entry = ttk.Entry(frame, width=35)
        self.desc_entry.grid(row=2, column=1, pady=10)

        self.error_label = ttk.Label(frame, text="", bootstyle=DANGER, font=('Microsoft YaHei', 9))
        self.error_label.grid(row=3, column=0, columnspan=2, sticky='w', pady=(10, 0), padx=10)

        if reminder:
            self.time_entry.insert(0, reminder.get('time', ''))
            self.title_entry.insert(0, reminder.get('title', ''))
            self.desc_entry.insert(0, reminder.get('description', ''))

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="确定", command=self.ok, bootstyle=PRIMARY, width=12).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel, bootstyle=SECONDARY, width=12).pack(side='left', padx=10)

        self.dialog.wait_window()

    def ok(self):
        time_val = self.time_entry.get().strip()
        title = self.title_entry.get().strip()
        desc = self.desc_entry.get().strip()

        if not time_val:
            self.error_label.config(text="⚠ 请输入时间")
            return
        if not re.match(r'^([01]\d|2[0-3]):([0-5]\d)$', time_val):
            self.error_label.config(text="⚠ 时间格式错误，请使用 HH:MM 格式（如 09:00）")
            return
        if not title:
            self.error_label.config(text="⚠ 请输入标题")
            return

        self.result = {'time': time_val, 'title': title, 'description': desc}
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


class CourseDialog:
    def __init__(self, parent, title, course=None):
        self.result = None

        self.dialog = ttk.Toplevel(parent, title=title)
        self.dialog.geometry("500x350")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog, padding=20)
        frame.pack(fill='both', expand=True)

        fields = [('课程名称:', 'name'), ('时间:', 'time'), ('地点:', 'location'), ('教师:', 'teacher'), ('备注:', 'note')]
        self.entries = {}

        for i, (label, key) in enumerate(fields):
            ttk.Label(frame, text=label, font=('Microsoft YaHei', 10)).grid(row=i, column=0, sticky='e', pady=8, padx=10)
            entry = ttk.Entry(frame, width=35)
            entry.grid(row=i, column=1, pady=8)
            self.entries[key] = entry

        self.error_label = ttk.Label(frame, text="", bootstyle=DANGER, font=('Microsoft YaHei', 9))
        self.error_label.grid(row=len(fields), column=0, columnspan=2, sticky='w', pady=(10, 0), padx=10)

        if course:
            self.entries['name'].insert(0, course.get('name', ''))
            self.entries['time'].insert(0, course.get('time', ''))
            self.entries['location'].insert(0, course.get('location', ''))
            self.entries['teacher'].insert(0, course.get('teacher', ''))
            self.entries['note'].insert(0, course.get('note', ''))

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="确定", command=self.ok, bootstyle=PRIMARY, width=12).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel, bootstyle=SECONDARY, width=12).pack(side='left', padx=10)

        self.dialog.wait_window()

    def ok(self):
        name = self.entries['name'].get().strip()
        time_val = self.entries['time'].get().strip()

        if not name:
            self.error_label.config(text="⚠ 请输入课程名称")
            return
        if not time_val:
            self.error_label.config(text="⚠ 请输入上课时间")
            return

        self.result = {
            'name': name,
            'time': time_val,
            'location': self.entries['location'].get().strip(),
            'teacher': self.entries['teacher'].get().strip(),
            'note': self.entries['note'].get().strip()
        }
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


class TimeSlotDialog:
    def __init__(self, parent, config_manager):
        self.saved = False
        self.config_manager = config_manager

        self.dialog = ttk.Toplevel(parent, title="编辑上课时间段")
        self.dialog.geometry("550x500")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog, padding=20)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text="自定义各节课的上课时间段", font=('Microsoft YaHei', 12, 'bold')).pack(anchor='w', pady=(0, 10))

        time_slots = config_manager.get_time_slots()
        self.entries = []

        list_frame = ttk.Frame(frame)
        list_frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(list_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        ttk.Label(scroll_frame, text="节次", font=('Microsoft YaHei', 9, 'bold'), width=10).grid(row=0, column=0, padx=5, pady=5)
        ttk.Label(scroll_frame, text="开始时间", font=('Microsoft YaHei', 9, 'bold'), width=12).grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(scroll_frame, text="结束时间", font=('Microsoft YaHei', 9, 'bold'), width=12).grid(row=0, column=2, padx=5, pady=5)

        for i, slot in enumerate(time_slots):
            row = i + 1
            label_entry = ttk.Entry(scroll_frame, width=12)
            label_entry.insert(0, slot['label'])
            label_entry.grid(row=row, column=0, padx=5, pady=4)
            start_entry = ttk.Entry(scroll_frame, width=12)
            start_entry.insert(0, slot['start'])
            start_entry.grid(row=row, column=1, padx=5, pady=4)
            end_entry = ttk.Entry(scroll_frame, width=12)
            end_entry.insert(0, slot['end'])
            end_entry.grid(row=row, column=2, padx=5, pady=4)
            self.entries.append((label_entry, start_entry, end_entry))

        btn_row = len(time_slots) + 1
        ttk.Button(scroll_frame, text="➕ 添加一行", command=lambda: self._add_row(scroll_frame, btn_row)).grid(row=btn_row, column=0, columnspan=3, pady=10)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        hint = ttk.Label(frame, text="提示：格式为 08:00（24小时制），可自行增加或删除时间段", font=('Microsoft YaHei', 9), bootstyle=SECONDARY)
        hint.pack(pady=(10, 0))

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="💾 保存", command=self.save, bootstyle=PRIMARY, width=12).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="↩️ 恢复默认", command=self.reset_default, bootstyle=WARNING, width=12).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel, bootstyle=SECONDARY, width=12).pack(side='left', padx=10)

        self.dialog.wait_window()

    def _add_row(self, parent_frame, row):
        label_entry = ttk.Entry(parent_frame, width=12)
        label_entry.grid(row=row, column=0, padx=5, pady=4)
        start_entry = ttk.Entry(parent_frame, width=12)
        start_entry.grid(row=row, column=1, padx=5, pady=4)
        end_entry = ttk.Entry(parent_frame, width=12)
        end_entry.grid(row=row, column=2, padx=5, pady=4)
        self.entries.append((label_entry, start_entry, end_entry))

    def save(self):
        time_slots = []
        for label_entry, start_entry, end_entry in self.entries:
            label = label_entry.get().strip()
            start = start_entry.get().strip()
            end = end_entry.get().strip()
            if label and start and end:
                time_slots.append({"label": label, "start": start, "end": end})
        if not time_slots:
            messagebox.showwarning("提示", "至少需要保留一个时间段")
            return
        self.config_manager.save_time_slots(time_slots)
        self.saved = True
        self.dialog.destroy()

    def reset_default(self):
        self.config_manager.save_time_slots(DEFAULT_TIME_SLOTS)
        self.saved = True
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


class _TaskProgressDialog:
    def __init__(self, root, app_dir, gui, task_mode, status_msg, success_msg, is_async=False):
        self.root = root
        self.app_dir = app_dir
        self.gui = gui
        self.task_mode = task_mode
        self.status_msg = status_msg
        self.success_msg = success_msg
        self.is_async = is_async
        self.process = None
        self.cancelled = False
        self.output_issues = []
        self.started_at = time.time()
        self.last_output_at = self.started_at
        self.heartbeat_job = None

        self.dialog = ttk.Toplevel(root, title=status_msg)
        self.dialog.geometry("650x450")
        self.dialog.transient(root)
        self.dialog.protocol("WM_DELETE_WINDOW", self.cancel)
        self.dialog.lift()
        self.dialog.focus_force()
        self.dialog.attributes("-topmost", True)
        self.dialog.after(1000, lambda: self.dialog.attributes("-topmost", False))

        frame = ttk.Frame(self.dialog, padding=15)
        frame.pack(fill='both', expand=True)

        header = ttk.Frame(frame)
        header.pack(fill='x', pady=(0, 10))
        ttk.Label(header, text="📋 任务执行中...", font=('Microsoft YaHei', 12, 'bold')).pack(side='left')
        self.progress = ttk.Progressbar(header, mode='indeterminate', length=200)
        self.progress.pack(side='right', padx=(10, 0))
        self.progress.start(10)

        text_frame = ttk.Frame(frame)
        text_frame.pack(fill='both', expand=True)

        self.text = tk.Text(text_frame, wrap='word', height=18, font=('Consolas', 10),
                             bg='#1e1e1e', fg='#d4d4d4', insertbackground='white',
                             state='disabled', relief='flat', borderwidth=0)
        text_scroll = ttk.Scrollbar(text_frame, orient='vertical', command=self.text.yview)
        self.text.configure(yscrollcommand=text_scroll.set)
        self.text.pack(side='left', fill='both', expand=True)
        text_scroll.pack(side='right', fill='y')

        self.text.tag_config('info', foreground='#569cd6')
        self.text.tag_config('success', foreground='#6a9955')
        self.text.tag_config('error', foreground='#f44747')
        self.text.tag_config('warning', foreground='#dcdcaa')

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=(10, 0))
        self.cancel_btn = ttk.Button(btn_frame, text="❌ 取消", command=self.cancel, bootstyle=DANGER)
        self.cancel_btn.pack(side='right')

    def _append_text(self, text, tag=None):
        try:
            self.last_output_at = time.time()
            self.text.configure(state='normal')
            if tag:
                self.text.insert('end', text + '\n', tag)
            else:
                self.text.insert('end', text + '\n')
            self.text.see('end')
            self.text.configure(state='disabled')
            self.dialog.update()
        except tk.TclError:
            pass

    def _start_heartbeat(self):
        def tick():
            if self.cancelled:
                return
            if self.process and self.process.poll() is None:
                elapsed = int(time.time() - self.started_at)
                idle = int(time.time() - self.last_output_at)
                if idle >= 15:
                    self._append_text(f"仍在运行，已用时 {elapsed} 秒。网络抓取可能需要几分钟，请不要重复点击。", 'info')
                self.heartbeat_job = self.dialog.after(15000, tick)

        self.heartbeat_job = self.dialog.after(15000, tick)

    def _log_to_file(self, msg):
        try:
            log_dir = self.app_dir / "logs"
            log_dir.mkdir(exist_ok=True)
            with open(log_dir / "gui_debug.log", 'a', encoding='utf-8') as f:
                f.write(f"[{datetime.now().isoformat()}] {msg}\n")
        except Exception:
            pass

    def _remember_issue(self, line, tag):
        if tag not in {'error', 'warning'}:
            return
        clean = line.strip()
        if clean and clean not in self.output_issues:
            self.output_issues.append(clean)

    def _issue_summary(self):
        if not self.output_issues:
            return ""
        head = self.output_issues[0]
        if len(self.output_issues) == 1:
            return head
        return f"{head}；另有 {len(self.output_issues) - 1} 条问题，请查看运行窗口。"

    def cancel(self):
        self.cancelled = True
        if self.process and self.process.poll() is None:
            self.process.terminate()
            try:
                self.process.wait(timeout=5)
            except Exception:
                self.process.kill()
        self._append_text("", 'info')
        self._append_text("⛔ 用户取消操作", 'warning')
        self._append_text("", 'info')
        self._show_done_button()

    def _show_done_button(self):
        self.progress.stop()
        if self.heartbeat_job:
            try:
                self.dialog.after_cancel(self.heartbeat_job)
            except tk.TclError:
                pass
            self.heartbeat_job = None
        self.cancel_btn.configure(text="✅ 关闭", command=self.dialog.destroy)

    def run(self):
        self._append_text(self.status_msg, 'info')
        self._append_text("任务窗口已打开。后端启动后会在这里持续输出日志。", 'info')
        self._append_text("如果 15 秒内没有新日志，会自动显示仍在运行的提示。", 'info')
        self._append_text("", 'info')

        def task_thread():
            try:
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = subprocess.SW_HIDE

                if getattr(sys, 'frozen', False):
                    exe_path = sys.executable
                    cmd = [exe_path, "--task", self.task_mode] if self.task_mode != "all" else [exe_path, "--task"]
                else:
                    script_path = _get_project_root() / "daily_assistant.py"
                    cmd = [sys.executable, str(script_path), self.task_mode]

                if self.is_async:
                    cmd.append("--async")

                self._log_to_file(f"任务启动: {' '.join(cmd)}")
                self.root.after(0, lambda: self._append_text(f"$ {' '.join(cmd)}", 'info'))
                self.root.after(0, lambda: self._append_text("", 'info'))

                self.process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace",
                    cwd=str(self.app_dir),
                    startupinfo=startupinfo,
                    creationflags=subprocess.CREATE_NO_WINDOW
                )
                self.root.after(0, self._start_heartbeat)

                for line in iter(self.process.stdout.readline, ''):
                    if self.cancelled:
                        break
                    if line:
                        line_stripped = line.rstrip('\n\r')
                        self._log_to_file(line_stripped)
                        tag = None
                        if any(w in line_stripped for w in ['成功', '完成', '✅']):
                            tag = 'success'
                        elif any(w in line_stripped for w in ['ERROR', '错误', '异常', '失败', 'failed', 'Failed', '❌']):
                            tag = 'error'
                        elif any(w in line_stripped for w in ['警告', 'WARNING', '⚠']):
                            tag = 'warning'
                        self._remember_issue(line_stripped, tag)
                        self.root.after(0, lambda t=line_stripped, tg=tag: self._append_text(t, tg))

                returncode = self.process.wait()

                self._log_to_file(f"返回码: {returncode}")
                self.root.after(0, self._show_done_button)

                if returncode == 0 and not self.cancelled:
                    self.root.after(0, lambda: self._append_text("", 'info'))
                    self.root.after(0, lambda: self._append_text("✅ " + self.success_msg, 'success'))
                    self.root.after(0, lambda: self.gui.status_var.set(self.success_msg))
                    self.root.after(0, self.gui.refresh_home_status)
                    if self.output_issues:
                        summary = self._issue_summary()
                        self.root.after(0, lambda s=summary: self._append_text("⚠ 部分流程存在问题：" + s, 'warning'))
                        self.root.after(100, lambda s=summary: self.gui.notify_task_issue("部分流程未完成", s, severity="warning"))
                    else:
                        self.root.after(1500, self.dialog.destroy)
                elif not self.cancelled:
                    self.root.after(0, lambda: self._append_text("", 'info'))
                    self.root.after(0, lambda: self._append_text("❌ 任务执行失败", 'error'))
                    self.root.after(0, lambda: self.gui.status_var.set("运行失败"))
                    message = self._issue_summary() or f"任务返回码：{returncode}"
                    self.root.after(100, lambda m=message: self.gui.notify_task_issue("任务执行失败", m, severity="error"))

            except Exception as e:
                import traceback
                self._log_to_file(f"异常: {traceback.format_exc()}")
                self.root.after(0, lambda: self._append_text(f"❌ 异常: {e}", 'error'))
                self.root.after(0, self._show_done_button)
                self.root.after(0, lambda: self.gui.status_var.set("运行失败"))
                self.root.after(100, lambda err=str(e): self.gui.notify_task_issue("任务执行异常", err, severity="error"))

        self.dialog.grab_set()
        thread = threading.Thread(target=task_thread, daemon=True)
        thread.start()
        self.dialog.wait_window()


# ========== 主程序入口 ==========

def run_task_mode():
    if getattr(sys, 'frozen', False):
        source_dir = Path(sys.executable).parent
        app_dir = source_dir
        sys.path.insert(0, str(source_dir / "_internal"))
    else:
        source_dir = _get_source_dir()
        app_dir = _get_app_dir()
        sys.path.insert(0, str(source_dir))

    log_dir = app_dir / "logs"
    log_dir.mkdir(exist_ok=True)
    with open(log_dir / "task_mode.log", 'a', encoding='utf-8') as f:
        f.write(f"[{datetime.now().isoformat()}] 任务模式启动" + "\n")
        f.write(f"app_dir: {app_dir}" + "\n")
        f.write(f"sys.argv: {sys.argv}" + "\n")

    from . import daily_assistant

    is_async = "--async" in sys.argv

    mode = "all"
    for i, arg in enumerate(sys.argv):
        if i > 0 and not arg.startswith("--") and sys.argv[i-1] == "--task":
            mode = arg
            break

    if is_async:
        asyncio.run(daily_assistant.main_async(mode=mode))
    else:
        daily_assistant.main(mode=mode)


def main():
    if not TTKBOOTSTRAP_AVAILABLE:
        raise RuntimeError(
            "GUI dependency missing: ttkbootstrap. Install dependencies with "
            "`python -m pip install -r requirements.txt`, or run backend task mode "
            "with `python gui_app.py --task remind|crawl|all`."
        )
    # 使用 ttkbootstrap 初始化现代化窗口主题，'litera' 是一个非常干净、护眼的光亮主题
    root = ttk.Window(title="📚 Daily Automation - 学术自动化助手", themename="litera", size=(1000, 750))
    app = DailyAutomationApp(root)
    root.mainloop()


if __name__ == "__main__":
    app_dir = _get_app_dir()
    log_dir = app_dir / "logs"
    log_dir.mkdir(exist_ok=True)
    with open(log_dir / "startup.log", 'a', encoding='utf-8') as f:
        f.write(f"[{datetime.now().isoformat()}] 程序启动" + "\n")
        f.write(f"sys.argv: {sys.argv}" + "\n")
        f.write(f"frozen: {getattr(sys, 'frozen', False)}" + "\n")
        f.write(f"app_dir: {app_dir}" + "\n")

    if len(sys.argv) > 1 and sys.argv[1] == "--task":
        try:
            run_task_mode()
        except Exception as e:
            import traceback
            with open(log_dir / "startup.log", 'a', encoding='utf-8') as f:
                f.write(f"FATAL: {traceback.format_exc()}" + "\n")
            error_root = tk.Tk()
            error_root.withdraw()
            error_root.lift()
            error_root.attributes('-topmost', True)
            messagebox.showerror("严重错误", f"任务模式启动失败:\n{e}\n\n详细信息已记录到 logs/startup.log")
            error_root.destroy()
    else:
        main()
