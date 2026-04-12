#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import json
import csv
import re
import threading
import subprocess
import sys
import webbrowser
from pathlib import Path
from datetime import datetime
from io import StringIO

from config_manager import ConfigManager, DEFAULT_NEWS_SOURCES, SOURCE_CATEGORIES


POPULAR_SCHOOLS = [
    "浙江大学", "清华大学", "北京大学", "复旦大学", "上海交通大学",
    "南京大学", "中国科学技术大学", "武汉大学", "华中科技大学", "中山大学",
    "同济大学", "东南大学", "哈尔滨工业大学", "西安交通大学", "北京航空航天大学",
    "四川大学", "南开大学", "天津大学", "厦门大学", "中国人民大学",
    "其他（手动输入）"
]

GRADES = ["大一", "大二", "大三", "大四", "研一", "研二", "研三", "博一", "博二", "博三"]

INTEREST_AREAS = [
    ("🤖 人工智能", ["artificial intelligence", "machine learning", "deep learning", "NLP", "computer vision"], ["人工智能", "机器学习", "深度学习"]),
    ("📊 数据科学", ["data science", "big data", "data mining", "statistics", "data analysis"], ["数据科学", "大数据", "数据挖掘"]),
    ("💻 计算机系统", ["operating systems", "distributed systems", "cloud computing", "database"], ["计算机系统", "分布式系统", "云计算"]),
    ("🔒 网络安全", ["cybersecurity", "cryptography", "network security", "privacy"], ["网络安全", "密码学"]),
    ("💰 经济学", ["economics", "econometrics", "macroeconomics", "microeconomics"], ["经济学", "计量经济学"]),
    ("📈 金融学", ["finance", "financial engineering", "quantitative finance", "fintech"], ["金融学", "金融工程"]),
    ("🏛️ 公共管理", ["public governance", "public policy", "digital governance", "e-government"], ["公共管理", "数字治理"]),
    ("🧠 心理学", ["psychology", "cognitive science", "behavioral science"], ["心理学", "认知科学"]),
    ("🏥 医学", ["medicine", "biomedical", "clinical research", "public health"], ["医学", "生物医学"]),
    ("📚 教育学", ["education", "pedagogy", "educational technology"], ["教育学", "教育技术"]),
    ("⚖️ 法学", ["law", "jurisprudence", "constitutional law"], ["法学", "法律"]),
    ("✍️ 文学", ["literature", "linguistics", "cultural studies"], ["文学", "语言学"]),
    ("🌱 环境科学", ["environmental science", "climate change", "sustainability"], ["环境科学", "气候变化"]),
    ("🔬 材料科学", ["materials science", "nanotechnology", "polymer science"], ["材料科学", "纳米技术"]),
]

DAY_NAMES_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
DAY_NAMES_EN = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DAY_MAP = dict(zip(DAY_NAMES_CN, DAY_NAMES_EN))


def _get_app_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    else:
        return Path(__file__).parent


def _decode_pwd(value: str) -> str:
    import base64
    if not value:
        return ""
    if value.startswith("fernet:"):
        try:
            from password_crypto import PasswordCrypto
            crypto = PasswordCrypto(_get_app_dir())
            return crypto.decrypt(value)
        except Exception:
            return value
    if value.startswith("enc:"):
        try:
            return base64.b64decode(value[4:]).decode('utf-8')
        except Exception:
            return value
    return value


def _encode_pwd(value: str) -> str:
    import base64
    if value and not value.startswith("enc:"):
        return "enc:" + base64.b64encode(value.encode('utf-8')).decode('utf-8')
    return value


def parse_course_table(text):
    courses = []
    lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
    if not lines:
        return courses

    separator = None
    for line in lines[:5]:
        if '\t' in line:
            separator = '\t'
            break
        elif '|' in line:
            separator = '|'
            break
        elif ',' in line and line.count(',') >= 2:
            separator = ','
            break

    rows = []
    for line in lines:
        if separator:
            parts = [p.strip() for p in line.split(separator)]
        else:
            parts = [p.strip() for p in re.split(r'\s{2,}', line)]
        if len(parts) >= 2:
            non_empty = sum(1 for p in parts if p)
            if non_empty <= 2 and len(parts) > 4:
                continue
            rows.append(parts)

    if not rows:
        return courses

    name_col = -1
    time_col = -1
    location_col = -1
    teacher_col = -1
    day_col = -1

    header_idx = -1
    for row_idx, row in enumerate(rows):
        temp_name = temp_time = temp_loc = temp_teacher = temp_day = -1
        found = False
        for i, cell in enumerate(row):
            cl = cell.lower()
            if any(k in cl for k in ['课程名称', '课程名', 'name', 'course name']):
                temp_name = i; found = True
            elif any(k in cl for k in ['上课时间', 'time', '节次']):
                temp_time = i; found = True
            elif any(k in cl for k in ['上课地点', 'location', 'room', '地点', '教室']):
                temp_loc = i; found = True
            elif any(k in cl for k in ['教师姓名', '任课教师', 'teacher', '授课教师']):
                temp_teacher = i; found = True
            elif any(k in cl for k in ['星期', 'day', 'week']) and not _looks_like_time_or_day(cell):
                temp_day = i; found = True
        if found and (temp_name >= 0 or temp_time >= 0):
            name_col, time_col, location_col, teacher_col, day_col = temp_name, temp_time, temp_loc, temp_teacher, temp_day
            header_idx = row_idx
            break

    if header_idx < 0:
        first_row = rows[0]
        for i, cell in enumerate(first_row):
            cl = cell.lower()
            if any(k in cl for k in ['课程名', '名称', 'course', 'name', '科目']):
                name_col = i
            elif any(k in cl for k in ['时间', 'time', '节次']):
                time_col = i
            elif any(k in cl for k in ['地点', '教室', 'location', 'room']):
                location_col = i
            elif any(k in cl for k in ['教师', 'teacher', '授课']):
                teacher_col = i
        if name_col < 0 and time_col < 0:
            if len(first_row) >= 4:
                if _looks_like_time_or_day(first_row[0]):
                    time_col, name_col, teacher_col, location_col = 0, 1, 2, 3
                else:
                    name_col, time_col, location_col, teacher_col = 0, 1, 2, 3
            elif len(first_row) >= 3:
                if _looks_like_time_or_day(first_row[0]):
                    time_col, name_col, location_col = 0, 1, 2
                else:
                    name_col, time_col, location_col = 0, 1, 2
            elif len(first_row) >= 2:
                if _looks_like_time_or_day(first_row[0]):
                    time_col, name_col = 0, 1
                else:
                    name_col, time_col = 0, 1
        header_idx = 0

    data_start = header_idx + 1

    for row in rows[data_start:]:
        if len(row) <= max(name_col, 0):
            continue

        name = row[name_col] if name_col >= 0 and name_col < len(row) else ""
        time_val = row[time_col] if time_col >= 0 and time_col < len(row) else ""
        location = row[location_col] if location_col >= 0 and location_col < len(row) else ""
        teacher = row[teacher_col] if teacher_col >= 0 and teacher_col < len(row) else ""
        day_str = row[day_col] if day_col >= 0 and day_col < len(row) else ""

        if not name:
            continue

        time_parts = [t.strip() for t in time_val.split(';') if t.strip()] if time_val else [""]
        loc_parts = [l.strip() for l in location.split(';') if l.strip()] if location else [""]

        for ti, tp in enumerate(time_parts):
            day = ""
            if day_str:
                day = _parse_day(day_str)
            if not day and tp:
                day = _parse_day(tp)
            if not day:
                day = "Monday"

            clean_time = _strip_day_from_time(tp) if tp and day else tp

            loc = loc_parts[ti] if ti < len(loc_parts) else (loc_parts[-1] if loc_parts else "")

            courses.append({
                'day': day,
                'name': name,
                'time': clean_time,
                'location': loc,
                'teacher': teacher,
                'note': ''
            })

    return courses


def _looks_like_time_or_day(text):
    if not text:
        return False
    if re.search(r'周[一二三四五六日末1-7]', text):
        return True
    if re.search(r'星期[一二三四五六日]', text):
        return True
    if re.search(r'(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)', text, re.IGNORECASE):
        return True
    if re.search(r'^[0-9]+-[0-9]+节', text):
        return True
    if re.search(r'^\d{1,2}:\d{2}', text):
        return True
    return False


def _parse_day(text):
    mapping = {
        '周一': 'Monday', '周二': 'Tuesday', '周三': 'Wednesday',
        '周四': 'Thursday', '周五': 'Friday', '周六': 'Saturday', '周日': 'Sunday',
        '星期一': 'Monday', '星期二': 'Tuesday', '星期三': 'Wednesday',
        '星期四': 'Thursday', '星期五': 'Friday', '星期六': 'Saturday', '星期日': 'Sunday',
        '周1': 'Monday', '周2': 'Tuesday', '周3': 'Wednesday',
        '周4': 'Thursday', '周5': 'Friday', '周6': 'Saturday', '周7': 'Sunday',
    }
    for cn, en in mapping.items():
        if cn in text:
            return en

    en_mapping = {
        'monday': 'Monday', 'tuesday': 'Tuesday', 'wednesday': 'Wednesday',
        'thursday': 'Thursday', 'friday': 'Friday', 'saturday': 'Saturday', 'sunday': 'Sunday',
    }
    text_lower = text.lower()
    for en_key, en_val in en_mapping.items():
        if en_key in text_lower:
            return en_val

    return ""


def _strip_day_from_time(text):
    text = re.sub(r'周[一二三四五六日末1-7]', '', text)
    text = re.sub(r'星期[一二三四五六日]', '', text)
    text = re.sub(r'(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)', '', text, flags=re.IGNORECASE)
    return text.strip()


class DailyAutomationApp:

    def __init__(self, root):
        self.root = root
        self.root.title("📚 Daily Automation - 学术自动化助手")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)

        self.app_dir = _get_app_dir()
        self.config_manager = ConfigManager(self.app_dir)

        is_first_run = self.config_manager.ensure_default_config()
        self.config_manager.ensure_default_schedule()

        self.create_ui()
        self.load_all_config()

        if is_first_run:
            self.root.after(100, self.show_setup_wizard)

    def show_setup_wizard(self):
        wizard = SetupWizard(self.root, self.config_manager)
        if wizard.completed:
            self.load_all_config()

    def create_ui(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)

        self.create_home_tab()
        self.create_sources_tab()
        self.create_keywords_tab()
        self.create_reminders_tab()
        self.create_email_tab()
        self.create_api_keys_tab()
        self.create_schedule_tab()
        self.create_about_tab()

        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief='sunken')
        status_bar.pack(fill='x', side='bottom')

    # ========== 首页 ==========
    def create_home_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="🏠 首页")

        title_frame = ttk.Frame(frame)
        title_frame.pack(fill='x', pady=20)
        ttk.Label(title_frame, text="📚 Daily Automation", font=('Microsoft YaHei', 24, 'bold')).pack()
        ttk.Label(title_frame, text="学术自动化助手 - 让信息找上门", font=('Microsoft YaHei', 12)).pack()

        card_frame = ttk.LabelFrame(frame, text="状态概览")
        card_frame.pack(fill='x', padx=20, pady=10)

        self.status_labels = {}
        stats = [('news_count', '新闻源', '0'), ('keyword_count', '关键词', '0'),
                 ('reminder_count', '提醒', '0'), ('email_status', '邮件', '未配置')]

        for i, (key, label, default) in enumerate(stats):
            col = ttk.Frame(card_frame)
            col.pack(side='left', expand=True, padx=20, pady=10)
            ttk.Label(col, text=label, font=('Microsoft YaHei', 10)).pack()
            self.status_labels[key] = ttk.Label(col, text=default, font=('Microsoft YaHei', 16, 'bold'))
            self.status_labels[key].pack()

        action_frame = ttk.LabelFrame(frame, text="快速操作")
        action_frame.pack(fill='x', padx=20, pady=10)

        btn_frame = ttk.Frame(action_frame)
        btn_frame.pack(pady=15)

        ttk.Button(btn_frame, text="📚 学术简报生成", command=self.run_crawl, width=18).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="⏰ 每日日程提醒", command=self.run_remind, width=18).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="📂 打开数据目录", command=self.open_data_dir, width=18).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="📤 导出配置", command=self.export_config, width=18).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="📥 导入配置", command=self.import_config, width=18).pack(side='left', padx=10)

        log_frame = ttk.LabelFrame(frame, text="最近运行日志")
        log_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, state='disabled')
        self.log_text.pack(fill='both', expand=True, padx=5, pady=5)

        self.refresh_home_status()

    def refresh_home_status(self):
        config = self.config_manager.get_config()

        sources = config.get('news_sources', [])
        enabled = len([s for s in sources if s.get('enabled', True)])
        self.status_labels['news_count'].config(text=f"{enabled}/{len(sources)}")

        keywords = config.get('keywords', []) + config.get('keywords_cn', [])
        self.status_labels['keyword_count'].config(text=str(len(keywords)))

        reminders = config.get('reminders', [])
        self.status_labels['reminder_count'].config(text=str(len(reminders)))

        email = config.get('email', {})
        self.status_labels['email_status'].config(
            text="已启用" if email.get('enabled') else "未启用"
        )

        self.load_recent_logs()

    def load_recent_logs(self):
        log_dir = self.app_dir / "logs"
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)

        if log_dir.exists():
            log_files = sorted(log_dir.glob("*.log"), key=lambda x: x.name, reverse=True)[:3]
            for log_file in log_files:
                try:
                    with open(log_file, 'r', encoding='utf-8') as f:
                        lines = f.readlines()[-20:]
                        self.log_text.insert(tk.END, f"=== {log_file.name} ===" + "\n")
                        self.log_text.insert(tk.END, ''.join(lines) + "\n\n")
                except Exception:
                    pass
        else:
            self.log_text.insert(tk.END, "暂无运行记录")

        self.log_text.config(state='disabled')

    # ========== 新闻源管理 ==========
    def create_sources_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="📰 新闻源")

        list_frame = ttk.LabelFrame(frame, text="已配置的新闻源")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('name', 'url', 'type', 'enabled')
        self.sources_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=8)

        self.sources_tree.heading('name', text='名称')
        self.sources_tree.heading('url', text='URL')
        self.sources_tree.heading('type', text='类型')
        self.sources_tree.heading('enabled', text='状态')

        self.sources_tree.column('name', width=120)
        self.sources_tree.column('url', width=350)
        self.sources_tree.column('type', width=80)
        self.sources_tree.column('enabled', width=80)

        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.sources_tree.yview)
        self.sources_tree.configure(yscrollcommand=scrollbar.set)

        self.sources_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(btn_frame, text="➕ 添加", command=self.add_source).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✏️ 编辑", command=self.edit_source).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🗑️ 删除", command=self.delete_source).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🔄 切换状态", command=self.toggle_source).pack(side='left', padx=5)

    def load_sources(self):
        for item in self.sources_tree.get_children():
            self.sources_tree.delete(item)

        sources = self.config_manager.get_news_sources()
        for source in sources:
            enabled = "✅ 启用" if source.get('enabled', True) else "❌ 禁用"
            self.sources_tree.insert('', 'end', values=(
                source.get('name', ''),
                source.get('url', ''),
                source.get('type', 'rss'),
                enabled
            ))

    def add_source(self):
        dialog = SourceDialog(self.root, "添加新闻源")
        if dialog.result:
            self.config_manager.add_news_source(**dialog.result)
            self.load_sources()
            self.refresh_home_status()
            messagebox.showinfo("成功", "新闻源已添加")

    def edit_source(self):
        selected = self.sources_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要编辑的新闻源")
            return

        index = self.sources_tree.index(selected[0])
        sources = self.config_manager.get_news_sources()
        source = sources[index]

        dialog = SourceDialog(self.root, "编辑新闻源", source)
        if dialog.result:
            self.config_manager.update_news_source(index, **dialog.result)
            self.load_sources()
            messagebox.showinfo("成功", "新闻源已更新")

    def delete_source(self):
        selected = self.sources_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的新闻源")
            return

        if messagebox.askyesno("确认", "确定要删除此新闻源吗？"):
            index = self.sources_tree.index(selected[0])
            self.config_manager.delete_news_source(index)
            self.load_sources()
            self.refresh_home_status()

    def toggle_source(self):
        selected = self.sources_tree.selection()
        if not selected:
            return

        index = self.sources_tree.index(selected[0])
        sources = self.config_manager.get_news_sources()
        current = sources[index].get('enabled', True)
        self.config_manager.update_news_source(index, enabled=not current)
        self.load_sources()

    # ========== 关键词管理 ==========
    def create_keywords_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="🔑 关键词")

        ttk.Label(frame, text="设置用于筛选学术信息的关键词（每行一个）", font=('Microsoft YaHei', 10)).pack(pady=10)

        col_frame = ttk.Frame(frame)
        col_frame.pack(fill='both', expand=True, padx=10, pady=5)

        en_frame = ttk.LabelFrame(col_frame, text="英文关键词")
        en_frame.pack(side='left', fill='both', expand=True, padx=5)

        self.en_keywords_text = scrolledtext.ScrolledText(en_frame, width=40)
        self.en_keywords_text.pack(fill='both', expand=True, padx=5, pady=5)

        cn_frame = ttk.LabelFrame(col_frame, text="中文关键词")
        cn_frame.pack(side='left', fill='both', expand=True, padx=5)

        self.cn_keywords_text = scrolledtext.ScrolledText(cn_frame, width=40)
        self.cn_keywords_text.pack(fill='both', expand=True, padx=5, pady=5)

        ttk.Button(frame, text="💾 保存关键词", command=self.save_keywords).pack(pady=10)

        suggest_frame = ttk.LabelFrame(frame, text="💡 关键词建议")
        suggest_frame.pack(fill='x', padx=10, pady=5)

        suggestions = "学术研究: artificial intelligence, machine learning, deep learning\n" \
                     "公共管理: public governance, digital governance, e-government\n" \
                     "数据科学: big data, data mining, algorithm"
        ttk.Label(suggest_frame, text=suggestions, justify='left').pack(padx=10, pady=5)

    def load_keywords(self):
        keywords = self.config_manager.get_keywords()

        self.en_keywords_text.delete(1.0, tk.END)
        self.en_keywords_text.insert(tk.END, '\n'.join(keywords.get('keywords', [])))

        self.cn_keywords_text.delete(1.0, tk.END)
        self.cn_keywords_text.insert(tk.END, '\n'.join(keywords.get('keywords_cn', [])))

    def save_keywords(self):
        en_text = self.en_keywords_text.get(1.0, tk.END)
        cn_text = self.cn_keywords_text.get(1.0, tk.END)

        en_keywords = [k.strip() for k in en_text.split('\n') if k.strip()]
        cn_keywords = [k.strip() for k in cn_text.split('\n') if k.strip()]

        if self.config_manager.update_keywords(en_keywords, cn_keywords):
            messagebox.showinfo("成功", "关键词已保存")
            self.refresh_home_status()

    # ========== 提醒管理 ==========
    def create_reminders_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="⏰ 提醒")

        list_frame = ttk.LabelFrame(frame, text="已配置的提醒")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('time', 'title', 'description')
        self.reminders_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=8)

        self.reminders_tree.heading('time', text='时间')
        self.reminders_tree.heading('title', text='标题')
        self.reminders_tree.heading('description', text='描述')

        self.reminders_tree.column('time', width=80)
        self.reminders_tree.column('title', width=150)
        self.reminders_tree.column('description', width=400)

        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.reminders_tree.yview)
        self.reminders_tree.configure(yscrollcommand=scrollbar.set)

        self.reminders_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(btn_frame, text="➕ 添加", command=self.add_reminder).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✏️ 编辑", command=self.edit_reminder).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🗑️ 删除", command=self.delete_reminder).pack(side='left', padx=5)

    def load_reminders(self):
        for item in self.reminders_tree.get_children():
            self.reminders_tree.delete(item)

        reminders = self.config_manager.get_reminders()
        for r in reminders:
            self.reminders_tree.insert('', 'end', values=(
                r.get('time', ''),
                r.get('title', ''),
                r.get('description', '')
            ))

    def add_reminder(self):
        dialog = ReminderDialog(self.root, "添加提醒")
        if dialog.result:
            self.config_manager.add_reminder(**dialog.result)
            self.load_reminders()
            self.refresh_home_status()
            messagebox.showinfo("成功", "提醒已添加")

    def edit_reminder(self):
        selected = self.reminders_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要编辑的提醒")
            return

        index = self.reminders_tree.index(selected[0])
        reminders = self.config_manager.get_reminders()
        reminder = reminders[index]

        dialog = ReminderDialog(self.root, "编辑提醒", reminder)
        if dialog.result:
            self.config_manager.update_reminder(index, **dialog.result)
            self.load_reminders()
            messagebox.showinfo("成功", "提醒已更新")

    def delete_reminder(self):
        selected = self.reminders_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的提醒")
            return

        if messagebox.askyesno("确认", "确定要删除此提醒吗？"):
            index = self.reminders_tree.index(selected[0])
            self.config_manager.delete_reminder(index)
            self.load_reminders()
            self.refresh_home_status()

    # ========== 邮箱配置 ==========
    def create_email_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="📧 邮箱")

        self.email_enabled = tk.BooleanVar()
        ttk.Checkbutton(frame, text="启用邮件发送", variable=self.email_enabled).pack(pady=10)

        form_frame = ttk.Frame(frame)
        form_frame.pack(fill='x', padx=20, pady=10)

        labels = ['SMTP服务器:', 'SMTP端口:', '发件人邮箱:', '授权密码:', '收件人邮箱:', '邮件主题前缀:']
        self.email_entries = {}

        for i, label in enumerate(labels):
            row = ttk.Frame(form_frame)
            row.pack(fill='x', pady=5)

            ttk.Label(row, text=label, width=15).pack(side='left')

            key = label.replace(':', '').replace(' ', '_')
            entry = ttk.Entry(row, width=50)
            if '密码' in label:
                entry.config(show='*')
            entry.pack(side='left', padx=5)
            self.email_entries[key] = entry

        ttk.Button(frame, text="💾 保存邮箱配置", command=self.save_email_config).pack(pady=15)

        help_frame = ttk.LabelFrame(frame, text="📋 常用邮箱SMTP配置")
        help_frame.pack(fill='x', padx=20, pady=10)

        help_text = """
QQ邮箱: smtp.qq.com 端口587 (需要开启SMTP服务并获取授权码)
163邮箱: smtp.163.com 端口25 (需要开启SMTP服务)
Gmail: smtp.gmail.com 端口587 (需要应用专用密码)
Outlook: smtp-mail.outlook.com 端口587
        """
        ttk.Label(help_frame, text=help_text, justify='left').pack(padx=10, pady=5)

    def load_email_config(self):
        email = self.config_manager.get_email_config()
        self.email_enabled.set(email.get('enabled', False))
        for key in self.email_entries:
            self.email_entries[key].delete(0, tk.END)
        self.email_entries['SMTP服务器'].insert(0, email.get('smtp_server', 'smtp.qq.com'))
        self.email_entries['SMTP端口'].insert(0, str(email.get('smtp_port', 587)))
        self.email_entries['发件人邮箱'].insert(0, email.get('sender_email', ''))
        self.email_entries['授权密码'].insert(0, _decode_pwd(email.get('sender_password', '')))
        self.email_entries['收件人邮箱'].insert(0, email.get('receiver_email', ''))
        self.email_entries['邮件主题前缀'].insert(0, email.get('subject_prefix', '[学术简报]'))

    def save_email_config(self):
        port_str = self.email_entries['SMTP端口'].get().strip()
        try:
            port = int(port_str) if port_str else 587
        except ValueError:
            messagebox.showerror("错误", "SMTP端口必须为数字")
            return

        sender_email = self.email_entries['发件人邮箱'].get().strip()
        sender_pwd = self.email_entries['授权密码'].get().strip()
        receiver_email = self.email_entries['收件人邮箱'].get().strip()

        auto_enabled = False
        if sender_email and sender_pwd and receiver_email and not self.email_enabled.get():
            self.email_enabled.set(True)
            auto_enabled = True

        config = {
            'enabled': self.email_enabled.get(),
            'smtp_server': self.email_entries['SMTP服务器'].get(),
            'smtp_port': port,
            'sender_email': sender_email,
            'sender_password': _encode_pwd(sender_pwd),
            'receiver_email': receiver_email,
            'subject_prefix': self.email_entries['邮件主题前缀'].get()
        }

        if self.config_manager.update_email_config(config):
            msg = "邮箱配置已保存"
            if auto_enabled:
                msg += "\n\n✅ 已自动启用邮件发送功能"
            messagebox.showinfo("成功", msg)
            self.refresh_home_status()

    # ========== API密钥配置 ==========
    def create_api_keys_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="🔑 API密钥")

        ttk.Label(frame, text="学术API密钥配置",
                  font=('Microsoft YaHei', 14, 'bold')).pack(pady=10)
        ttk.Label(frame, text="配置API密钥可提升访问速率和配额，部分API必须配置密钥才能使用",
                  foreground='gray').pack(pady=5)

        form_frame = ttk.Frame(frame)
        form_frame.pack(fill='x', padx=20, pady=10)

        self.api_key_entries = {}
        api_key_config = [
            ('semantic_scholar', 'Semantic Scholar:', '可选 - 无Key限流100次/5分钟'),
            ('openalex', 'OpenAlex:', '可选 - 有Key可提升速率'),
            ('core', 'CORE:', '必须 - 无Key无法使用'),
        ]

        for i, (key_name, label, hint) in enumerate(api_key_config):
            row = ttk.Frame(form_frame)
            row.pack(fill='x', pady=8)

            ttk.Label(row, text=label, width=18).pack(side='left')

            entry = ttk.Entry(row, width=50, show='*')
            entry.pack(side='left', padx=5)
            self.api_key_entries[key_name] = entry

            ttk.Label(row, text=hint, foreground='gray').pack(side='left', padx=5)

        self.show_api_keys = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame, text="显示密钥", variable=self.show_api_keys,
                        command=self._toggle_api_key_visibility).pack(pady=5)

        ttk.Button(frame, text="💾 保存API密钥配置",
                   command=self.save_api_keys_config).pack(pady=10)

        help_frame = ttk.LabelFrame(frame, text="🔗 获取API密钥")
        help_frame.pack(fill='x', padx=20, pady=10)

        links = [
            ("Semantic Scholar API Key", "https://www.semanticscholar.org/product/api#api-key"),
            ("OpenAlex API Key", "https://docs.openalex.org/how-to-use-the-api/get-an-api-key"),
            ("CORE API Key", "https://core.ac.uk/services/api"),
        ]

        for name, url in links:
            link_frame = ttk.Frame(help_frame)
            link_frame.pack(fill='x', padx=10, pady=3)
            ttk.Label(link_frame, text=f"• {name}:").pack(side='left')
            link_label = ttk.Label(link_frame, text=url, foreground='blue', cursor='hand2')
            link_label.pack(side='left', padx=5)
            link_label.bind('<Button-1>', lambda e, u=url: webbrowser.open(u))

        self.api_key_status = ttk.Label(frame, text="", foreground='orange')
        self.api_key_status.pack(pady=5)

    def _toggle_api_key_visibility(self):
        show = self.show_api_keys.get()
        for entry in self.api_key_entries.values():
            entry.config(show='' if show else '*')

    def load_api_keys_config(self):
        if not hasattr(self, 'api_key_entries'):
            return
        keys = self.config_manager.get_api_keys()
        for key_name, entry in self.api_key_entries.items():
            entry.delete(0, tk.END)
            value = keys.get(key_name, '')
            if value and value.startswith('enc:'):
                value = _decode_pwd(value)
            entry.insert(0, value)

        self._check_api_key_status()

    def save_api_keys_config(self):
        api_keys = {}
        for key_name, entry in self.api_key_entries.items():
            value = entry.get().strip()
            if value and not value.startswith('enc:'):
                value = _encode_pwd(value)
            api_keys[key_name] = value

        if self.config_manager.update_api_keys(api_keys):
            messagebox.showinfo("成功", "API密钥配置已保存")
            self._check_api_key_status()
            self.refresh_home_status()

    def _check_api_key_status(self):
        if not hasattr(self, 'api_key_entries'):
            return
        missing = self.config_manager.check_missing_api_keys()
        if missing:
            required_missing = [m for m in missing if m['required']]
            optional_missing = [m for m in missing if not m['required']]
            msg_parts = []
            if required_missing:
                names = ', '.join(m['name'] for m in required_missing)
                msg_parts.append(f"⚠️ 必须配置: {names}")
            if optional_missing:
                names = ', '.join(m['name'] for m in optional_missing)
                msg_parts.append(f"💡 建议配置: {names}")
            self.api_key_status.config(text=' | '.join(msg_parts))
        else:
            self.api_key_status.config(text="✅ 所有API密钥已配置", foreground='green')

    # ========== 课程表 ==========
    def create_schedule_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="📅 课程表")

        select_frame = ttk.Frame(frame)
        select_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(select_frame, text="选择日期:").pack(side='left')
        self.day_var = tk.StringVar()
        self.day_combo = ttk.Combobox(select_frame, textvariable=self.day_var, values=DAY_NAMES_CN, width=10, state='readonly')
        self.day_combo.current(0)
        self.day_combo.pack(side='left', padx=10)
        self.day_combo.bind('<<ComboboxSelected>>', self.on_day_change)

        list_frame = ttk.LabelFrame(frame, text="当日课程")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('name', 'time', 'location', 'teacher')
        self.schedule_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=10)

        self.schedule_tree.heading('name', text='课程名称')
        self.schedule_tree.heading('time', text='时间')
        self.schedule_tree.heading('location', text='地点')
        self.schedule_tree.heading('teacher', text='教师')

        self.schedule_tree.column('name', width=200)
        self.schedule_tree.column('time', width=120)
        self.schedule_tree.column('location', width=200)
        self.schedule_tree.column('teacher', width=100)

        self.schedule_tree.pack(fill='both', expand=True)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(btn_frame, text="➕ 添加课程", command=self.add_course).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🗑️ 删除课程", command=self.delete_course).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📥 导入课表", command=self.import_schedule).pack(side='left', padx=5)

    def load_schedule(self):
        self.on_day_change(None)

    def on_day_change(self, event):
        for item in self.schedule_tree.get_children():
            self.schedule_tree.delete(item)

        day_name = self.day_var.get()
        day = DAY_MAP.get(day_name, 'Monday')

        schedule = self.config_manager.get_week_schedule()
        courses = schedule.get(day, [])

        for course in courses:
            self.schedule_tree.insert('', 'end', values=(
                course.get('name', ''),
                course.get('time', ''),
                course.get('location', ''),
                course.get('teacher', '')
            ))

    def add_course(self):
        day_name = self.day_var.get()
        day = DAY_MAP.get(day_name, 'Monday')

        dialog = CourseDialog(self.root, "添加课程")
        if dialog.result:
            self.config_manager.add_course(day, **dialog.result)
            self.on_day_change(None)
            messagebox.showinfo("成功", "课程已添加")

    def delete_course(self):
        selected = self.schedule_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的课程")
            return

        if messagebox.askyesno("确认", "确定要删除此课程吗？"):
            day_name = self.day_var.get()
            day = DAY_MAP.get(day_name, 'Monday')
            index = self.schedule_tree.index(selected[0])
            self.config_manager.delete_course(day, index)
            self.on_day_change(None)

    def import_schedule(self):
        dialog = ScheduleImportDialog(self.root, self.config_manager)
        if dialog.imported:
            self.on_day_change(None)
            self.refresh_home_status()

    # ========== 关于 ==========
    def create_about_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="ℹ️ 关于")

        ttk.Label(frame, text="📚 Daily Automation", font=('Microsoft YaHei', 20, 'bold')).pack(pady=30)
        ttk.Label(frame, text="学术自动化助手 v1.0", font=('Microsoft YaHei', 12)).pack()
        ttk.Label(frame, text="让信息找上门，省下时间做重要的事", font=('Microsoft YaHei', 10)).pack(pady=10)

        ttk.Separator(frame, orient='horizontal').pack(fill='x', padx=50, pady=20)

        features = ttk.LabelFrame(frame, text="功能特性")
        features.pack(fill='x', padx=50, pady=10)

        feature_text = """
✅ 多源学术信息爬取（arXiv、Semantic Scholar等）
✅ 智能关键词筛选
✅ 每日简报生成
✅ 邮件自动发送
✅ 日程提醒
✅ 课程表管理
        """
        ttk.Label(features, text=feature_text, justify='left').pack(padx=10, pady=5)

        ttk.Button(frame, text="🔄 重置为默认配置", command=self.reset_config).pack(pady=20)

    def reset_config(self):
        if messagebox.askyesno("确认", "确定要重置为默认配置吗？所有自定义设置将丢失。"):
            self.config_manager.reset_to_default()
            self.load_all_config()
            messagebox.showinfo("成功", "已重置为默认配置")

    # ========== 通用方法 ==========
    def load_all_config(self):
        self.load_sources()
        self.load_keywords()
        self.load_reminders()
        self.load_email_config()
        self.load_api_keys_config()
        self.load_schedule()
        self.refresh_home_status()

    def _run_subprocess_task(self, task_mode: str, status_msg: str, success_msg: str):
        self.status_var.set(status_msg)
        self.root.update()

        def run_task():
            try:
                log_dir = self.app_dir / "logs"
                log_dir.mkdir(exist_ok=True)

                debug_log = log_dir / "gui_debug.log"
                def log(msg):
                    with open(debug_log, 'a', encoding='utf-8') as f:
                        f.write(f"[{datetime.now().isoformat()}] {msg}" + "\n")

                log("=== 任务开始 ===")
                log(f"app_dir: {self.app_dir}")

                if getattr(sys, 'frozen', False):
                    exe_path = sys.executable
                    log(f"打包模式，运行: {exe_path} --task {task_mode}")

                    startupinfo = subprocess.STARTUPINFO()
                    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                    startupinfo.wShowWindow = subprocess.SW_HIDE

                    cmd = [exe_path, "--task", task_mode] if task_mode != "all" else [exe_path, "--task"]
                    result = subprocess.run(
                        cmd,
                        capture_output=True, text=True, encoding="utf-8", errors="replace", cwd=str(self.app_dir),
                        timeout=600,
                        startupinfo=startupinfo,
                        creationflags=subprocess.CREATE_NO_WINDOW
                    )
                else:
                    script_path = self.app_dir / "daily_assistant.py"
                    log(f"开发模式，运行: {script_path} {task_mode}")
                    result = subprocess.run(
                        [sys.executable, str(script_path), task_mode],
                        capture_output=True, text=True, encoding="utf-8", errors="replace",
                        cwd=str(self.app_dir),
                        timeout=600
                    )

                log(f"返回码: {result.returncode}")
                if result.stdout:
                    log(f"输出: {result.stdout[:500]}")
                if result.stderr:
                    log(f"错误: {result.stderr[:500]}")

                if result.returncode == 0:
                    log("任务执行成功")
                    self.root.after(0, lambda: self.status_var.set(success_msg))
                    self.root.after(0, self.refresh_home_status)
                else:
                    log(f"任务执行失败: {result.returncode}")
                    self.root.after(0, lambda: messagebox.showerror("运行错误", "任务执行失败" + "\n" + f"{result.stderr}"))
                    self.root.after(0, lambda: self.status_var.set("运行失败"))

                log("=== 任务结束 ===")

            except subprocess.TimeoutExpired:
                log("任务超时")
                self.root.after(0, lambda: messagebox.showerror("错误", "任务执行超时"))
                self.root.after(0, lambda: self.status_var.set("运行失败"))
            except Exception as e:
                import traceback
                log(f"异常: {traceback.format_exc()}")
                self.root.after(0, lambda: messagebox.showerror("错误", "执行失败:" + "\n" + f"{str(e)}"))
                self.root.after(0, lambda: self.status_var.set("运行失败"))

        thread = threading.Thread(target=run_task, daemon=True)
        thread.start()

    def run_once(self):
        self._run_subprocess_task("all", "正在运行...", "运行完成")

    def run_crawl(self):
        self._run_subprocess_task("crawl", "正在生成学术简报...", "学术简报生成完成")

    def run_remind(self):
        self._run_subprocess_task("remind", "正在检查日程提醒...", "日程提醒检查完成")

    def open_data_dir(self):
        data_dir = self.app_dir / "data"
        data_dir.mkdir(exist_ok=True)
        webbrowser.open(str(data_dir))

    def export_config(self):
        data = self.config_manager.export_all_config()
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")],
            initialfile=f"daily_automation_config_{datetime.now().strftime('%Y%m%d')}.json"
        )
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("成功", f"配置已导出到: {filename}")

    def import_config(self):
        filename = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")]
        )
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if self.config_manager.import_all_config(data):
                    self.load_all_config()
                    messagebox.showinfo("成功", "配置已导入")
            except Exception as e:
                messagebox.showerror("错误", f"导入失败: {e}")


# ========== 首次启动向导 ==========

class SetupWizard:
    def __init__(self, parent, config_manager):
        self.config_manager = config_manager
        self.completed = False
        self.current_step = 0
        self.total_steps = 4
        self.step_titles = ["基本信息", "课表导入", "数据源选择", "兴趣方向"]
        self.selected_interests = []

        self.wizard = tk.Toplevel(parent)
        self.wizard.title("欢迎使用 Daily Automation")
        self.wizard.geometry("700x600")
        self.wizard.resizable(True, True)
        self.wizard.transient(parent)
        self.wizard.grab_set()
        self.wizard.protocol("WM_DELETE_WINDOW", self._on_skip)

        header = ttk.Frame(self.wizard)
        header.pack(fill='x', padx=20, pady=(15, 5))
        ttk.Label(header, text="📚 Daily Automation 初始设置",
                  font=('Microsoft YaHei', 16, 'bold')).pack()
        ttk.Label(header, text="快速完成配置，3步即可开始使用", foreground='gray').pack()

        self.step_indicator = ttk.Frame(self.wizard)
        self.step_indicator.pack(fill='x', padx=20, pady=5)
        self.step_labels = []
        for i, title in enumerate(self.step_titles):
            lbl = ttk.Label(self.step_indicator, text=f"  {i+1}. {title}  ", font=('Microsoft YaHei', 9))
            lbl.pack(side='left', padx=5)
            self.step_labels.append(lbl)

        ttk.Separator(self.wizard, orient='horizontal').pack(fill='x', padx=20, pady=5)

        self.content_frame = ttk.Frame(self.wizard)
        self.content_frame.pack(fill='both', expand=True, padx=20, pady=5)

        btn_frame = ttk.Frame(self.wizard)
        btn_frame.pack(fill='x', padx=20, pady=15)

        self.skip_btn = ttk.Button(btn_frame, text="跳过全部", command=self._on_skip)
        self.skip_btn.pack(side='left')

        self.back_btn = ttk.Button(btn_frame, text="◀ 上一步", command=self._prev_step, state='disabled')
        self.back_btn.pack(side='right', padx=5)

        self.next_btn = ttk.Button(btn_frame, text="下一步 ▶", command=self._next_step)
        self.next_btn.pack(side='right', padx=5)

        self._build_step(0)
        self._update_step_indicator()
        self.wizard.wait_window()

    def _build_step(self, step):
        for w in self.content_frame.winfo_children():
            w.destroy()

        if step == 0:
            self._build_step_basic()
        elif step == 1:
            self._build_step_schedule()
        elif step == 2:
            self._build_step_sources()
        elif step == 3:
            self._build_step_interests()

    def _build_step_basic(self):
        f = self.content_frame

        ttk.Label(f, text="📝 基本信息设置", font=('Microsoft YaHei', 13, 'bold')).pack(anchor='w', pady=(5, 15))

        row1 = ttk.Frame(f)
        row1.pack(fill='x', pady=8)
        ttk.Label(row1, text="学校:", width=10).pack(side='left')
        self.school_var = tk.StringVar()
        self.school_combo = ttk.Combobox(row1, textvariable=self.school_var, values=POPULAR_SCHOOLS, width=30)
        self.school_combo.pack(side='left', padx=5)
        self.school_combo.bind('<<ComboboxSelected>>', self._on_school_select)

        self.school_entry = ttk.Entry(row1, width=30)
        self.school_entry.pack(side='left', padx=5)
        self.school_entry.pack_forget()

        row2 = ttk.Frame(f)
        row2.pack(fill='x', pady=8)
        ttk.Label(row2, text="年级:", width=10).pack(side='left')
        self.grade_var = tk.StringVar()
        ttk.Combobox(row2, textvariable=self.grade_var, values=GRADES, width=28, state='readonly').pack(side='left', padx=5)

        row3 = ttk.Frame(f)
        row3.pack(fill='x', pady=8)
        ttk.Label(row3, text="学期:", width=10).pack(side='left')
        self.semester_var = tk.StringVar()
        now = datetime.now()
        year = now.year
        sem = "春夏" if now.month >= 2 and now.month <= 7 else "秋冬"
        self.semester_var.set(f"{year-1}-{year}学年{sem}学期")
        ttk.Entry(row3, textvariable=self.semester_var, width=30).pack(side='left', padx=5)

        ttk.Label(f, text="💡 这些信息仅用于个性化你的每日计划，不会上传到任何服务器",
                  foreground='gray', font=('Microsoft YaHei', 9)).pack(anchor='w', pady=(20, 0))

    def _on_school_select(self, event):
        if self.school_var.get() == "其他（手动输入）":
            self.school_combo.pack_forget()
            self.school_entry.pack(side='left', padx=5)
            self.school_entry.focus_set()
        else:
            self.school_entry.pack_forget()
            self.school_combo.pack(side='left', padx=5)

    def _build_step_schedule(self):
        f = self.content_frame

        ttk.Label(f, text="📅 课表导入", font=('Microsoft YaHei', 13, 'bold')).pack(anchor='w', pady=(5, 10))
        ttk.Label(f, text="从教务系统复制课表粘贴到下方，或选择CSV/Excel文件导入",
                  foreground='gray').pack(anchor='w', pady=(0, 10))

        method_frame = ttk.Frame(f)
        method_frame.pack(fill='x', pady=5)

        self.schedule_method = tk.StringVar(value='text')
        ttk.Radiobutton(method_frame, text="📝 粘贴文字", variable=self.schedule_method,
                        value='text', command=self._toggle_schedule_method).pack(side='left', padx=10)
        ttk.Radiobutton(method_frame, text="📂 导入文件", variable=self.schedule_method,
                        value='file', command=self._toggle_schedule_method).pack(side='left', padx=10)

        self.text_frame = ttk.Frame(f)
        self.text_frame.pack(fill='both', expand=True, pady=5)

        ttk.Label(self.text_frame, text="粘贴课表内容（支持从教务系统直接复制）:",
                  font=('Microsoft YaHei', 9)).pack(anchor='w')
        self.schedule_text = scrolledtext.ScrolledText(self.text_frame, height=8, width=60)
        self.schedule_text.pack(fill='both', expand=True, pady=3)
        ttk.Label(self.text_frame, text="支持格式：Tab分隔 / 逗号分隔 / 竖线分隔，含列名或不含均可",
                  foreground='gray', font=('Microsoft YaHei', 8)).pack(anchor='w')

        self.file_frame = ttk.Frame(f)

        self.file_path_var = tk.StringVar()
        ttk.Entry(self.file_frame, textvariable=self.file_path_var, width=45).pack(side='left', padx=5)
        ttk.Button(self.file_frame, text="浏览...", command=self._browse_schedule_file).pack(side='left', padx=5)
        ttk.Label(self.file_frame, text="支持 CSV / Excel (.xlsx) 文件", foreground='gray').pack(side='left', padx=5)

        self.preview_frame = ttk.LabelFrame(f, text="解析预览")
        self.preview_frame.pack(fill='both', expand=True, pady=5)

        self.preview_tree = ttk.Treeview(self.preview_frame, columns=('day', 'name', 'time', 'location', 'teacher'),
                                         show='headings', height=5)
        for col, heading, width in [('day', '星期', 60), ('name', '课程名', 120), ('time', '时间', 100),
                                     ('location', '地点', 100), ('teacher', '教师', 80)]:
            self.preview_tree.heading(col, text=heading)
            self.preview_tree.column(col, width=width)
        self.preview_tree.pack(fill='both', expand=True)

        btn_row = ttk.Frame(f)
        btn_row.pack(fill='x', pady=5)
        ttk.Button(btn_row, text="🔍 解析预览", command=self._preview_schedule).pack(side='left', padx=5)

    def _toggle_schedule_method(self):
        if self.schedule_method.get() == 'text':
            self.file_frame.pack_forget()
            self.text_frame.pack(fill='both', expand=True, pady=5)
        else:
            self.text_frame.pack_forget()
            self.file_frame.pack(fill='x', pady=5)

    def _browse_schedule_file(self):
        filetypes = [("表格文件", "*.csv *.xlsx *.xls"), ("CSV文件", "*.csv"), ("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            self.file_path_var.set(filename)

    def _preview_schedule(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)

        courses = self._get_parsed_courses()
        day_cn_map = {v: k for k, v in DAY_MAP.items()}
        for c in courses:
            day_cn = day_cn_map.get(c['day'], c['day'])
            self.preview_tree.insert('', 'end', values=(day_cn, c['name'], c['time'], c['location'], c['teacher']))

        if not courses:
            messagebox.showinfo("提示", "未能解析出课程信息，请检查输入格式")

    def _get_parsed_courses(self):
        if self.schedule_method.get() == 'text':
            text = self.schedule_text.get(1.0, tk.END).strip()
            if not text:
                return []
            return parse_course_table(text)
        else:
            filepath = self.file_path_var.get().strip()
            if not filepath:
                return []
            return self._parse_file(filepath)

    def _parse_file(self, filepath):
        path = Path(filepath)
        if not path.exists():
            return []

        try:
            if path.suffix.lower() == '.csv':
                return self._parse_csv_file(path)
            elif path.suffix.lower() in ('.xlsx', '.xls'):
                return self._parse_excel_file(path)
            else:
                with open(path, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
                return parse_course_table(content)
        except Exception as e:
            messagebox.showerror("解析错误", f"文件解析失败: {e}")
            return []

    def _parse_csv_file(self, path):
        courses = []
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            return courses

        text = '\n'.join(['\t'.join(row) for row in rows])
        return parse_course_table(text)

    def _parse_excel_file(self, path):
        try:
            import openpyxl
        except ImportError:
            messagebox.showwarning("提示", "Excel文件需要安装 openpyxl 库。\n请运行: pip install openpyxl\n或改用CSV格式导入。")
            return []

        courses = []
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else '' for c in row]
            rows.append(cells)
        wb.close()

        if not rows:
            return courses

        text = '\n'.join(['\t'.join(row) for row in rows])
        return parse_course_table(text)

    def _build_step_sources(self):
        f = self.content_frame

        ttk.Label(f, text="数据源选择", font=('Microsoft YaHei', 13, 'bold')).pack(anchor='w', pady=(5, 5))
        ttk.Label(f, text="选择你关注的学术数据源（取消勾选可减少运行时间）", foreground='gray').pack(anchor='w', pady=(0, 10))

        canvas_frame = ttk.Frame(f)
        canvas_frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(canvas_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient='vertical', command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all('<MouseWheel>', _on_mousewheel)

        self.source_vars = {}
        current_category = None

        for source in DEFAULT_NEWS_SOURCES:
            cat_key = source.get('category', 'general_academic')
            cat_name = SOURCE_CATEGORIES.get(cat_key, cat_key)

            if cat_key != current_category:
                current_category = cat_key
                cat_label = ttk.Label(scroll_frame, text=f"  {cat_name}",
                                      font=('Microsoft YaHei', 10, 'bold'), foreground='#2c5f8a')
                cat_label.pack(anchor='w', padx=5, pady=(8, 2))
                ttk.Separator(scroll_frame, orient='horizontal').pack(fill='x', padx=5, pady=1)

            row_frame = ttk.Frame(scroll_frame)
            row_frame.pack(fill='x', padx=15, pady=1)

            var = tk.BooleanVar(value=source.get('enabled', True))
            self.source_vars[source['name']] = var

            cb = ttk.Checkbutton(row_frame, text=source['name'], variable=var)
            cb.pack(side='left')

            desc = source.get('desc', '')
            if desc:
                desc_label = ttk.Label(row_frame, text=f"  - {desc}",
                                       foreground='gray', font=('Microsoft YaHei', 8))
                desc_label.pack(side='left', padx=(5, 0))

            type_tag = source.get('type', 'web').upper()
            type_label = ttk.Label(row_frame, text=f"[{type_tag}]",
                                   foreground='#888', font=('Consolas', 8))
            type_label.pack(side='right', padx=5)

        btn_frame = ttk.Frame(f)
        btn_frame.pack(fill='x', pady=5)
        ttk.Button(btn_frame, text="全选", command=lambda: self._set_all_sources(True)).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="全不选", command=lambda: self._set_all_sources(False)).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="推荐配置", command=self._set_recommended_sources).pack(side='left', padx=5)

    def _set_all_sources(self, value):
        if hasattr(self, 'source_vars'):
            for var in self.source_vars.values():
                var.set(value)

    def _set_recommended_sources(self):
        if hasattr(self, 'source_vars'):
            for name, var in self.source_vars.items():
                default = next((s['enabled'] for s in DEFAULT_NEWS_SOURCES if s['name'] == name), True)
                var.set(default)

    def _build_step_interests(self):
        f = self.content_frame

        ttk.Label(f, text="🎯 兴趣方向", font=('Microsoft YaHei', 13, 'bold')).pack(anchor='w', pady=(5, 10))
        ttk.Label(f, text="选择你感兴趣的领域，将自动填充相关学术关键词（可多选）",
                  foreground='gray').pack(anchor='w', pady=(0, 10))

        tags_frame = ttk.Frame(f)
        tags_frame.pack(fill='x', pady=5)

        self.interest_vars = {}
        for i, (name, _, _) in enumerate(INTEREST_AREAS):
            row = i // 3
            col = i % 3
            var = tk.BooleanVar(value=False)
            self.interest_vars[name] = var
            cb = ttk.Checkbutton(tags_frame, text=name, variable=var)
            cb.grid(row=row, column=col, sticky='w', padx=10, pady=4)

        ttk.Separator(f, orient='horizontal').pack(fill='x', pady=10)

        custom_frame = ttk.LabelFrame(f, text="自定义关键词（可选）")
        custom_frame.pack(fill='both', expand=True, pady=5)

        dual = ttk.Frame(custom_frame)
        dual.pack(fill='both', expand=True, padx=5, pady=5)

        ttk.Label(dual, text="英文关键词（每行一个）:").pack(anchor='w')
        self.custom_en = scrolledtext.ScrolledText(dual, height=3, width=30)
        self.custom_en.pack(fill='both', expand=True, side='left', padx=(0, 5))

        ttk.Label(dual, text="中文关键词（每行一个）:").pack(anchor='w')
        self.custom_cn = scrolledtext.ScrolledText(dual, height=3, width=30)
        self.custom_cn.pack(fill='both', expand=True, side='left')

    def _update_step_indicator(self):
        for i, lbl in enumerate(self.step_labels):
            if i < self.current_step:
                lbl.config(foreground='green')
            elif i == self.current_step:
                lbl.config(foreground='blue', font=('Microsoft YaHei', 10, 'bold'))
            else:
                lbl.config(foreground='gray')

    def _next_step(self):
        if self.current_step == 1:
            courses = self._get_parsed_courses()
            if courses:
                schedule = self.config_manager.get_schedule()
                if 'week_schedule' not in schedule:
                    schedule['week_schedule'] = {}
                for day in DAY_NAMES_EN:
                    if day not in schedule['week_schedule']:
                        schedule['week_schedule'][day] = []

                for c in courses:
                    day = c['day']
                    if day not in schedule['week_schedule']:
                        schedule['week_schedule'][day] = []
                    schedule['week_schedule'][day].append({
                        'name': c['name'],
                        'time': c['time'],
                        'location': c['location'],
                        'teacher': c['teacher'],
                        'note': c.get('note', '')
                    })
                self.config_manager.save_schedule(schedule)

        if self.current_step == 2:
            self._save_source_selection()

        if self.current_step < self.total_steps - 1:
            self.current_step += 1
            self._build_step(self.current_step)
            self._update_step_indicator()
            self.back_btn.config(state='normal')
            if self.current_step == self.total_steps - 1:
                self.next_btn.config(text="完成")
        else:
            self._save_basic_info()
            self._save_interests()
            self.completed = True
            self.wizard.destroy()

    def _prev_step(self):
        if self.current_step > 0:
            self.current_step -= 1
            self._build_step(self.current_step)
            self._update_step_indicator()
            self.back_btn.config(state='disabled' if self.current_step == 0 else 'normal')
            self.next_btn.config(text="下一步 ▶")

    def _on_skip(self):
        self._save_basic_info()
        self.completed = True
        self.wizard.destroy()

    def _save_basic_info(self):
        if not hasattr(self, 'semester_var'):
            return

        semester = self.semester_var.get().strip() if hasattr(self, 'semester_var') else ""
        if semester:
            schedule = self.config_manager.get_schedule()
            schedule['semester'] = semester
            self.config_manager.save_schedule(schedule)

        config = self.config_manager.get_config()
        if not config.get('reminders'):
            config['reminders'] = [
                {"time": "09:00", "title": "学术早报", "description": "今日最新学术资讯已整理完毕，请查阅。"},
                {"time": "22:00", "title": "晚间复盘", "description": "今日工作告一段落，回顾一下收获。"}
            ]
            self.config_manager.save_config(config)

    def _save_source_selection(self):
        if not hasattr(self, 'source_vars'):
            return

        config = self.config_manager.get_config()
        sources = config.get('news_sources', [])

        for source in sources:
            if source['name'] in self.source_vars:
                source['enabled'] = self.source_vars[source['name']].get()

        config['news_sources'] = sources
        self.config_manager.save_config(config)

    def _save_interests(self):
        if not hasattr(self, 'interest_vars'):
            return

        en_keywords = []
        cn_keywords = []

        for name, en_list, cn_list in INTEREST_AREAS:
            if self.interest_vars.get(name, tk.BooleanVar()).get():
                en_keywords.extend(en_list)
                cn_keywords.extend(cn_list)

        if hasattr(self, 'custom_en'):
            custom_en = [k.strip() for k in self.custom_en.get(1.0, tk.END).split('\n') if k.strip()]
            custom_cn = [k.strip() for k in self.custom_cn.get(1.0, tk.END).split('\n') if k.strip()]
            en_keywords.extend(custom_en)
            cn_keywords.extend(custom_cn)

        en_keywords = list(dict.fromkeys(en_keywords))
        cn_keywords = list(dict.fromkeys(cn_keywords))

        if en_keywords or cn_keywords:
            config = self.config_manager.get_config()
            existing_en = config.get('keywords', [])
            existing_cn = config.get('keywords_cn', [])
            merged_en = list(dict.fromkeys(existing_en + en_keywords))
            merged_cn = list(dict.fromkeys(existing_cn + cn_keywords))
            self.config_manager.update_keywords(merged_en, merged_cn)


# ========== 课表导入对话框 ==========

class ScheduleImportDialog:
    def __init__(self, parent, config_manager):
        self.config_manager = config_manager
        self.imported = False

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("📥 导入课表")
        self.dialog.geometry("700x550")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        ttk.Label(self.dialog, text="导入课表", font=('Microsoft YaHei', 14, 'bold')).pack(pady=(10, 5))
        ttk.Label(self.dialog, text="从教务系统复制课表粘贴到下方，或选择CSV/Excel文件导入",
                  foreground='gray').pack(pady=(0, 10))

        method_frame = ttk.Frame(self.dialog)
        method_frame.pack(fill='x', padx=15, pady=5)

        self.method = tk.StringVar(value='text')
        ttk.Radiobutton(method_frame, text="📝 粘贴文字", variable=self.method,
                        value='text', command=self._toggle_method).pack(side='left', padx=10)
        ttk.Radiobutton(method_frame, text="📂 导入文件", variable=self.method,
                        value='file', command=self._toggle_method).pack(side='left', padx=10)

        self.text_frame = ttk.Frame(self.dialog)
        self.text_frame.pack(fill='both', expand=True, padx=15, pady=5)

        ttk.Label(self.text_frame, text="粘贴课表内容（支持从教务系统直接复制）:").pack(anchor='w')
        self.text_area = scrolledtext.ScrolledText(self.text_frame, height=6, width=70)
        self.text_area.pack(fill='both', expand=True, pady=3)
        ttk.Label(self.text_frame, text="支持格式：Tab分隔 / 逗号分隔 / 竖线分隔",
                  foreground='gray', font=('Microsoft YaHei', 8)).pack(anchor='w')

        self.file_frame = ttk.Frame(self.dialog)
        self.file_path = tk.StringVar()
        ttk.Entry(self.file_frame, textvariable=self.file_path, width=50).pack(side='left', padx=5)
        ttk.Button(self.file_frame, text="浏览...", command=self._browse_file).pack(side='left', padx=5)
        ttk.Label(self.file_frame, text="支持 CSV / Excel (.xlsx)", foreground='gray').pack(side='left', padx=5)

        preview_lf = ttk.LabelFrame(self.dialog, text="解析预览")
        preview_lf.pack(fill='both', expand=True, padx=15, pady=5)

        self.preview_tree = ttk.Treeview(preview_lf, columns=('day', 'name', 'time', 'location', 'teacher'),
                                         show='headings', height=6)
        for col, heading, width in [('day', '星期', 60), ('name', '课程名', 140), ('time', '时间', 100),
                                     ('location', '地点', 120), ('teacher', '教师', 80)]:
            self.preview_tree.heading(col, text=heading)
            self.preview_tree.column(col, width=width)
        self.preview_tree.pack(fill='both', expand=True)

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(fill='x', padx=15, pady=10)

        ttk.Button(btn_frame, text="🔍 解析预览", command=self._preview).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✅ 确认导入", command=self._confirm).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="取消", command=self.dialog.destroy).pack(side='right', padx=5)

        self.parsed_courses = []
        self.dialog.wait_window()

    def _toggle_method(self):
        if self.method.get() == 'text':
            self.file_frame.pack_forget()
            self.text_frame.pack(fill='both', expand=True, padx=15, pady=5)
        else:
            self.text_frame.pack_forget()
            self.file_frame.pack(fill='x', padx=15, pady=5)

    def _browse_file(self):
        filetypes = [("表格文件", "*.csv *.xlsx *.xls"), ("CSV文件", "*.csv"), ("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            self.file_path.set(filename)

    def _preview(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)

        self.parsed_courses = self._parse()
        day_cn_map = {v: k for k, v in DAY_MAP.items()}
        for c in self.parsed_courses:
            day_cn = day_cn_map.get(c['day'], c['day'])
            self.preview_tree.insert('', 'end', values=(day_cn, c['name'], c['time'], c['location'], c['teacher']))

        if not self.parsed_courses:
            messagebox.showinfo("提示", "未能解析出课程信息，请检查输入格式")

    def _parse(self):
        if self.method.get() == 'text':
            text = self.text_area.get(1.0, tk.END).strip()
            if not text:
                return []
            return parse_course_table(text)
        else:
            filepath = self.file_path.get().strip()
            if not filepath:
                return []
            path = Path(filepath)
            if not path.exists():
                return []
            try:
                if path.suffix.lower() == '.csv':
                    return self._parse_csv(path)
                elif path.suffix.lower() in ('.xlsx', '.xls'):
                    return self._parse_excel(path)
                else:
                    with open(path, 'r', encoding='utf-8', errors='replace') as f:
                        return parse_course_table(f.read())
            except Exception as e:
                messagebox.showerror("解析错误", f"文件解析失败: {e}")
                return []

    def _parse_csv(self, path):
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            rows = list(csv.reader(f))
        if not rows:
            return []
        text = '\n'.join(['\t'.join(row) for row in rows])
        return parse_course_table(text)

    def _parse_excel(self, path):
        try:
            import openpyxl
        except ImportError:
            messagebox.showwarning("提示", "Excel文件需要安装 openpyxl 库。\n请运行: pip install openpyxl\n或改用CSV格式导入。")
            return []
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else '' for c in row]
            rows.append(cells)
        wb.close()
        if not rows:
            return []
        text = '\n'.join(['\t'.join(row) for row in rows])
        return parse_course_table(text)

    def _confirm(self):
        if not self.parsed_courses:
            self._preview()
            if not self.parsed_courses:
                return

        schedule = self.config_manager.get_schedule()
        if 'week_schedule' not in schedule:
            schedule['week_schedule'] = {}
        for day in DAY_NAMES_EN:
            if day not in schedule['week_schedule']:
                schedule['week_schedule'][day] = []

        for c in self.parsed_courses:
            day = c['day']
            if day not in schedule['week_schedule']:
                schedule['week_schedule'][day] = []
            schedule['week_schedule'][day].append({
                'name': c['name'],
                'time': c['time'],
                'location': c['location'],
                'teacher': c['teacher'],
                'note': c.get('note', '')
            })

        self.config_manager.save_schedule(schedule)
        self.imported = True
        messagebox.showinfo("成功", f"已导入 {len(self.parsed_courses)} 门课程")
        self.dialog.destroy()


# ========== 对话框类（带输入验证） ==========

class SourceDialog:
    def __init__(self, parent, title, source=None):
        self.result = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("520x230")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog)
        frame.pack(fill='both', expand=True, padx=20, pady=15)

        ttk.Label(frame, text="名称:").grid(row=0, column=0, sticky='e', pady=5)
        self.name_entry = ttk.Entry(frame, width=40)
        self.name_entry.grid(row=0, column=1, pady=5, sticky='w')

        ttk.Label(frame, text="URL:").grid(row=1, column=0, sticky='e', pady=5)
        self.url_entry = ttk.Entry(frame, width=40)
        self.url_entry.grid(row=1, column=1, pady=5, sticky='w')

        ttk.Label(frame, text="类型:").grid(row=2, column=0, sticky='e', pady=5)
        self.type_var = tk.StringVar(value='rss')
        ttk.Combobox(frame, textvariable=self.type_var, values=['rss', 'web', 'json'], width=37, state='readonly').grid(row=2, column=1, pady=5, sticky='w')

        self.error_label = ttk.Label(frame, text="", foreground='red', font=('Microsoft YaHei', 9))
        self.error_label.grid(row=3, column=0, columnspan=2, sticky='w', pady=(5, 0))

        if source:
            self.name_entry.insert(0, source.get('name', ''))
            self.url_entry.insert(0, source.get('url', ''))
            self.type_var.set(source.get('type', 'rss'))

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="确定", command=self.ok).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel).pack(side='left', padx=10)

        self.dialog.wait_window()

    def ok(self):
        name = self.name_entry.get().strip()
        url = self.url_entry.get().strip()

        if not name:
            self.error_label.config(text="⚠ 请输入名称")
            return
        if not url:
            self.error_label.config(text="⚠ 请输入URL")
            return
        if not url.startswith(('http://', 'https://')):
            self.error_label.config(text="⚠ URL必须以 http:// 或 https:// 开头")
            return

        self.result = {
            'name': name,
            'url': url,
            'source_type': self.type_var.get(),
            'enabled': True
        }
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


class ReminderDialog:
    def __init__(self, parent, title, reminder=None):
        self.result = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("470x220")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog)
        frame.pack(fill='both', expand=True, padx=20, pady=15)

        ttk.Label(frame, text="时间 (HH:MM):").grid(row=0, column=0, sticky='e', pady=5)
        self.time_entry = ttk.Entry(frame, width=30)
        self.time_entry.grid(row=0, column=1, pady=5)

        ttk.Label(frame, text="标题:").grid(row=1, column=0, sticky='e', pady=5)
        self.title_entry = ttk.Entry(frame, width=30)
        self.title_entry.grid(row=1, column=1, pady=5)

        ttk.Label(frame, text="描述:").grid(row=2, column=0, sticky='e', pady=5)
        self.desc_entry = ttk.Entry(frame, width=30)
        self.desc_entry.grid(row=2, column=1, pady=5)

        self.error_label = ttk.Label(frame, text="", foreground='red', font=('Microsoft YaHei', 9))
        self.error_label.grid(row=3, column=0, columnspan=2, sticky='w', pady=(5, 0))

        if reminder:
            self.time_entry.insert(0, reminder.get('time', ''))
            self.title_entry.insert(0, reminder.get('title', ''))
            self.desc_entry.insert(0, reminder.get('description', ''))

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="确定", command=self.ok).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel).pack(side='left', padx=10)

        self.dialog.wait_window()

    def ok(self):
        time_val = self.time_entry.get().strip()
        title = self.title_entry.get().strip()
        desc = self.desc_entry.get().strip()

        if not time_val:
            self.error_label.config(text="⚠ 请输入时间")
            return
        if not re.match(r'^([01]\d|2[0-3]):([0-5]\d)$', time_val):
            self.error_label.config(text="⚠ 时间格式错误，请使用 HH:MM 格式（如 09:00）")
            return
        if not title:
            self.error_label.config(text="⚠ 请输入标题")
            return

        self.result = {'time': time_val, 'title': title, 'description': desc}
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


class CourseDialog:
    def __init__(self, parent, title, course=None):
        self.result = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("450x280")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog)
        frame.pack(fill='both', expand=True, padx=20, pady=15)

        fields = [('课程名称:', 'name'), ('时间:', 'time'), ('地点:', 'location'), ('教师:', 'teacher'), ('备注:', 'note')]
        self.entries = {}

        for i, (label, key) in enumerate(fields):
            ttk.Label(frame, text=label).grid(row=i, column=0, sticky='e', pady=3)
            entry = ttk.Entry(frame, width=30)
            entry.grid(row=i, column=1, pady=3)
            self.entries[key] = entry

        self.error_label = ttk.Label(frame, text="", foreground='red', font=('Microsoft YaHei', 9))
        self.error_label.grid(row=len(fields), column=0, columnspan=2, sticky='w', pady=(5, 0))

        if course:
            self.entries['name'].insert(0, course.get('name', ''))
            self.entries['time'].insert(0, course.get('time', ''))
            self.entries['location'].insert(0, course.get('location', ''))
            self.entries['teacher'].insert(0, course.get('teacher', ''))
            self.entries['note'].insert(0, course.get('note', ''))

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="确定", command=self.ok).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel).pack(side='left', padx=10)

        self.dialog.wait_window()

    def ok(self):
        name = self.entries['name'].get().strip()
        time_val = self.entries['time'].get().strip()

        if not name:
            self.error_label.config(text="⚠ 请输入课程名称")
            return
        if not time_val:
            self.error_label.config(text="⚠ 请输入上课时间")
            return

        self.result = {
            'name': name,
            'time': time_val,
            'location': self.entries['location'].get().strip(),
            'teacher': self.entries['teacher'].get().strip(),
            'note': self.entries['note'].get().strip()
        }
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


# ========== 主程序入口 ==========

def run_task_mode():
    if getattr(sys, 'frozen', False):
        app_dir = Path(sys.executable).parent
    else:
        app_dir = Path(__file__).parent

    log_dir = app_dir / "logs"
    log_dir.mkdir(exist_ok=True)
    with open(log_dir / "task_mode.log", 'a', encoding='utf-8') as f:
        f.write(f"[{datetime.now().isoformat()}] 任务模式启动" + "\n")
        f.write(f"app_dir: {app_dir}" + "\n")
        f.write(f"sys.argv: {sys.argv}" + "\n")

    sys.path.insert(0, str(app_dir))
    import daily_assistant
    mode = sys.argv[2] if len(sys.argv) > 2 else "all"
    daily_assistant.main(mode=mode)


def main():
    root = tk.Tk()
    app = DailyAutomationApp(root)
    root.mainloop()


if __name__ == "__main__":
    app_dir = _get_app_dir()
    log_dir = app_dir / "logs"
    log_dir.mkdir(exist_ok=True)
    with open(log_dir / "startup.log", 'a', encoding='utf-8') as f:
        f.write(f"[{datetime.now().isoformat()}] 程序启动" + "\n")
        f.write(f"sys.argv: {sys.argv}" + "\n")
        f.write(f"frozen: {getattr(sys, 'frozen', False)}" + "\n")
        f.write(f"app_dir: {app_dir}" + "\n")

    if len(sys.argv) > 1 and sys.argv[1] == "--task":
        run_task_mode()
    else:
        main()
